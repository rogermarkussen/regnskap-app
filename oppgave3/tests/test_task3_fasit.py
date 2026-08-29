from __future__ import annotations

import sys
import unittest
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from monthly_close_data import _summary_frame  # noqa: E402
from shared.data_contract import load_data_contract  # noqa: E402


CONTRACT = load_data_contract(REPO_ROOT)
FASIT_PATH = CONTRACT.path("fasit.account_grouping_all")
FASIT_154301_PATH = CONTRACT.path("fasit.account_grouping_154301")
LEDGER_PATH = CONTRACT.path("common.ledger")
BUDGET_HEADER_PATH = CONTRACT.path("common.budget_header")
BUDGET_VALUE_PATH = CONTRACT.path("common.budget_values")
PUBLISHED_SUMMARY_PATH = CONTRACT.generated_dir("oppgave3") / "web" / "monthly_close_summary.parquet"
FASIT_PERIOD = "202603"
TOLERANCE_NOK = 0.02


def _fasit_total(workbook_path: Path, label: str) -> dict[str, float]:
    """Read a named total from Excel without relying on a fixed row number."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Kontogruppering med tall"]
        matches = [
            row
            for row in worksheet.iter_rows(values_only=True)
            if str(row[0] or "").strip() == label
        ]
    finally:
        workbook.close()
    if len(matches) != 1:
        raise AssertionError(
            f"Forventet én fasitrad med teksten {label!r}, fant {len(matches)}"
        )
    row = matches[0]
    return {
        "budsjett_nok": float(row[2]) * 1000,
        "hovedbok_nok": float(row[3]) * 1000,
        "avvik_nok": float(row[4]) * 1000,
    }


def _single_row(
    frame: pd.DataFrame, *, financing: str, category: str
) -> pd.Series:
    rows = frame[
        (frame["omfang"] == "Nkom")
        & (frame["finansiering"] == financing)
        & (frame["kategori"] == category)
        & (frame["kildestatus"] == "Beregnet")
    ]
    if len(rows) != 1:
        raise AssertionError(
            f"Forventet én Nkom-rad for {financing}/{category}, fant {len(rows)}"
        )
    return rows.iloc[0]


def _all_financing_total(frame: pd.DataFrame, category: str) -> pd.Series:
    """Aggregate ordinary Nkom rows; provisional 712 cash rows are excluded."""
    rows = frame[
        (frame["omfang"] == "Nkom")
        & (frame["kategori"] == category)
        & (frame["kildestatus"] == "Beregnet")
    ]
    if rows.empty:
        raise AssertionError(f"Fant ingen beregnede Nkom-rader for {category}")
    return rows[
        [
            "hovedbok_hittil_nok",
            "budsjett_hittil_nok",
            "avvik_hittil_nok",
        ]
    ].sum()


class Task3FasitTest(unittest.TestCase):
    """Reconcile task 3 calculations with the independent Excel answer key."""

    @classmethod
    def setUpClass(cls) -> None:
        required = [
            FASIT_PATH,
            FASIT_154301_PATH,
            LEDGER_PATH,
            BUDGET_HEADER_PATH,
            BUDGET_VALUE_PATH,
            PUBLISHED_SUMMARY_PATH,
        ]
        missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
        if missing:
            raise AssertionError(f"Mangler testgrunnlag: {', '.join(missing)}")

        connection = duckdb.connect()
        try:
            cls.calculated = _summary_frame(
                connection,
                LEDGER_PATH,
                BUDGET_HEADER_PATH,
                BUDGET_VALUE_PATH,
                FASIT_PERIOD,
            )
            cls.published = connection.execute(
                f"select * from read_parquet('{PUBLISHED_SUMMARY_PATH.as_posix()}')"
            ).df()
        finally:
            connection.close()

    def assertNokEqual(
        self, calculated: float, expected: float, message: str
    ) -> None:
        difference = float(calculated) - float(expected)
        self.assertAlmostEqual(
            float(calculated),
            float(expected),
            delta=TOLERANCE_NOK,
            msg=(
                f"{message}: beregnet {calculated:,.2f} NOK, "
                f"fasit {expected:,.2f} NOK, avvik {difference:,.2f} NOK"
            ),
        )

    def test_fasitens_periode_og_budsjettversjon(self) -> None:
        workbook = load_workbook(FASIT_PATH, read_only=True, data_only=True)
        try:
            worksheet = workbook["Kontogruppering med tall"]
            self.assertIn("01-03 2026", str(worksheet["A3"].value))
            self.assertIn("2026B", str(worksheet["A4"].value))
        finally:
            workbook.close()
        self.assertEqual(set(self.calculated["periode"]), {FASIT_PERIOD})
        self.assertEqual(set(self.calculated["budsjettversjon"]), {"2026B"})

    def test_154301_lonns_hovedbok_og_budsjett(self) -> None:
        calculated = _single_row(
            self.calculated, financing="154301", category="Lønnskostnader"
        )
        expected = _fasit_total(FASIT_154301_PATH, "Totale lønnskostnader")
        self.assertNokEqual(
            calculated["hovedbok_hittil_nok"],
            expected["hovedbok_nok"],
            "154301 lønn hovedbok",
        )
        self.assertNokEqual(
            calculated["budsjett_hittil_nok"],
            expected["budsjett_nok"],
            "154301 lønn budsjett",
        )

    def test_154301_adk_hovedbok_og_budsjett(self) -> None:
        calculated = _single_row(
            self.calculated, financing="154301", category="ADK"
        )
        expected = _fasit_total(
            FASIT_154301_PATH, "Totale andre driftskostnader"
        )
        self.assertNokEqual(
            calculated["hovedbok_hittil_nok"],
            expected["hovedbok_nok"],
            "154301 ADK hovedbok",
        )
        self.assertNokEqual(
            calculated["budsjett_hittil_nok"],
            expected["budsjett_nok"],
            "154301 ADK budsjett",
        )

    def test_154301_driftskostnader_hovedbok_mot_fasit(self) -> None:
        calculated = _single_row(
            self.calculated, financing="154301", category="Driftskostnader"
        )
        expected = _fasit_total(FASIT_154301_PATH, "Driftskostnader")
        self.assertNokEqual(
            calculated["hovedbok_hittil_nok"],
            expected["hovedbok_nok"],
            "154301 driftskostnader hovedbok",
        )

    def test_154301_driftskostnader_budsjett_mot_fasit(self) -> None:
        calculated = _single_row(
            self.calculated, financing="154301", category="Driftskostnader"
        )
        expected = _fasit_total(FASIT_154301_PATH, "Driftskostnader")
        self.assertNokEqual(
            calculated["budsjett_hittil_nok"],
            expected["budsjett_nok"],
            "154301 driftskostnader budsjett",
        )

    def test_154301_adk_avvik_har_fasitens_fortegn(self) -> None:
        calculated = _single_row(
            self.calculated, financing="154301", category="ADK"
        )
        expected = _fasit_total(
            FASIT_154301_PATH, "Totale andre driftskostnader"
        )
        self.assertNokEqual(
            calculated["avvik_hittil_nok"],
            expected["avvik_nok"],
            "154301 ADK avvik",
        )

    def test_nkom_lonns_hovedbok_mot_alle_fasit(self) -> None:
        calculated = _all_financing_total(self.calculated, "Lønnskostnader")
        expected = _fasit_total(FASIT_PATH, "Totale lønnskostnader")
        self.assertNokEqual(
            calculated["hovedbok_hittil_nok"],
            expected["hovedbok_nok"],
            "Nkom lønn hovedbok",
        )

    def test_nkom_lonns_budsjett_mot_alle_fasit(self) -> None:
        calculated = _all_financing_total(self.calculated, "Lønnskostnader")
        expected = _fasit_total(FASIT_PATH, "Totale lønnskostnader")
        self.assertNokEqual(
            calculated["budsjett_hittil_nok"],
            expected["budsjett_nok"],
            "Nkom lønn budsjett",
        )

    def test_nkom_adk_hovedbok_mot_alle_fasit(self) -> None:
        calculated = _all_financing_total(self.calculated, "ADK")
        expected = _fasit_total(FASIT_PATH, "Totale andre driftskostnader")
        self.assertNokEqual(
            calculated["hovedbok_hittil_nok"],
            expected["hovedbok_nok"],
            "Nkom ADK hovedbok",
        )

    def test_nkom_adk_budsjett_mot_alle_fasit(self) -> None:
        calculated = _all_financing_total(self.calculated, "ADK")
        expected = _fasit_total(FASIT_PATH, "Totale andre driftskostnader")
        self.assertNokEqual(
            calculated["budsjett_hittil_nok"],
            expected["budsjett_nok"],
            "Nkom ADK budsjett",
        )

    def test_nkom_driftskostnader_hovedbok_mot_alle_fasit(self) -> None:
        calculated = _all_financing_total(self.calculated, "Driftskostnader")
        expected = _fasit_total(FASIT_PATH, "Driftskostnader")
        self.assertNokEqual(
            calculated["hovedbok_hittil_nok"],
            expected["hovedbok_nok"],
            "Nkom driftskostnader hovedbok",
        )

    def test_nkom_driftskostnader_budsjett_mot_alle_fasit(self) -> None:
        calculated = _all_financing_total(self.calculated, "Driftskostnader")
        expected = _fasit_total(FASIT_PATH, "Driftskostnader")
        self.assertNokEqual(
            calculated["budsjett_hittil_nok"],
            expected["budsjett_nok"],
            "Nkom driftskostnader budsjett",
        )

    def test_publiserte_tall_er_reproduserbare(self) -> None:
        period = str(self.published["periode"].max())
        connection = duckdb.connect()
        try:
            recalculated = _summary_frame(
                connection,
                LEDGER_PATH,
                BUDGET_HEADER_PATH,
                BUDGET_VALUE_PATH,
                period,
            )
        finally:
            connection.close()

        keys = ["omfang", "omfang_id", "finansiering", "kategori"]
        measures = [
            "hovedbok_maaned_nok",
            "budsjett_maaned_nok",
            "avvik_maaned_nok",
            "hovedbok_hittil_nok",
            "budsjett_hittil_nok",
            "avvik_hittil_nok",
        ]
        published = self.published[
            (self.published["periode"] == period)
            & (self.published["kildestatus"] == "Beregnet")
        ]
        recalculated = recalculated[recalculated["kildestatus"] == "Beregnet"]
        comparison = published[keys + measures].merge(
            recalculated[keys + measures],
            on=keys,
            how="outer",
            suffixes=("_publisert", "_beregnet"),
            indicator=True,
        )
        self.assertTrue(
            (comparison["_merge"] == "both").all(),
            "Publisert og ny beregning har ulike rapportnøkler",
        )
        for measure in measures:
            difference = (
                comparison[f"{measure}_publisert"].fillna(0)
                - comparison[f"{measure}_beregnet"].fillna(0)
            ).abs()
            self.assertLessEqual(
                float(difference.max()),
                TOLERANCE_NOK,
                f"Publisert {measure} kan ikke reproduseres",
            )

    def test_publisert_avvik_er_budsjett_minus_hovedbok(self) -> None:
        for span in ("maaned", "forrige", "hittil"):
            columns = [
                f"hovedbok_{span}_nok",
                f"budsjett_{span}_nok",
                f"avvik_{span}_nok",
            ]
            rows = self.published.dropna(subset=columns)
            expected = rows[columns[1]] - rows[columns[0]]
            difference = (rows[columns[2]] - expected).abs()
            self.assertLessEqual(
                float(difference.max()),
                TOLERANCE_NOK,
                f"Publisert avvik for {span} er ikke budsjett minus hovedbok",
            )

    def test_publiserte_driftskostnader_inkluderer_avskrivninger(self) -> None:
        ordinary = self.published[self.published["kildestatus"] == "Beregnet"]
        keys = ["periode", "omfang", "omfang_id", "finansiering"]
        measures = [
            "hovedbok_maaned_nok",
            "budsjett_maaned_nok",
            "hovedbok_hittil_nok",
            "budsjett_hittil_nok",
        ]
        components = (
            ordinary[
                ordinary["kategori"].isin(
                    ["Lønnskostnader", "Avskrivninger", "ADK"]
                )
            ]
            .groupby(keys, as_index=False)[measures]
            .sum()
        )
        totals = ordinary[ordinary["kategori"] == "Driftskostnader"][
            keys + measures
        ]
        comparison = totals.merge(
            components, on=keys, suffixes=("_total", "_komponenter")
        )
        self.assertFalse(comparison.empty, "Fant ingen driftskostnadstotaler")
        for measure in measures:
            difference = (
                comparison[f"{measure}_total"]
                - comparison[f"{measure}_komponenter"]
            ).abs()
            self.assertLessEqual(
                float(difference.max()),
                TOLERANCE_NOK,
                f"{measure} summerer ikke lønn, avskrivninger og ADK",
            )

    def test_poster_uten_finansiering_er_synlige_i_nkom(self) -> None:
        rows = self.published[
            (self.published["omfang"] == "Nkom")
            & (self.published["finansiering"] == "Uten finansiering")
            & (self.published["kildestatus"] == "Beregnet")
        ]
        self.assertFalse(
            rows.empty,
            "Hovedboksposter uten finansiering mangler i Nkom-rapporten",
        )
        self.assertGreater(
            float(rows["hovedbok_hittil_nok"].abs().sum()),
            0,
            "Raden Uten finansiering inneholder ingen hovedbokstall",
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
