from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    from .project_data import task3_sources
    from .task3_rules import load_task3_rules
except ImportError:
    from project_data import task3_sources
    from task3_rules import load_task3_rules


RULES = load_task3_rules()
SECTIONS = RULES.sections
TEMPLATE_NAME = "Ønsket mal_mnds avsl.xlsx"
MONTH_NAMES = {
    1: "Januar",
    2: "Februar",
    3: "Mars",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


@dataclass(frozen=True)
class MonthlyCloseResult:
    summary: pd.DataFrame
    invoices: pd.DataFrame
    validations: pd.DataFrame
    period: str
    workbook_path: Path


def _budget_amount_sql(
    conn: duckdb.DuckDBPyConnection,
    path: Path,
    alias: str = "v",
) -> str:
    columns = {
        row[0]
        for row in conn.execute(
            f"describe select * from read_parquet('{path.as_posix()}')"
        ).fetchall()
    }
    candidates = [
        f"try_cast({alias}.{name} as double)"
        for name in ("amount", "amount1")
        if name in columns
    ]
    if not candidates:
        raise ValueError(f"{path.name} mangler både amount og amount1")
    return candidates[0] if len(candidates) == 1 else f"coalesce({', '.join(candidates)})"


def _category_case(alias: str) -> str:
    salary = RULES.account_categories["Lønnskostnader"]
    depreciation = RULES.account_categories["Avskrivninger"]
    adk = RULES.account_categories["ADK"]
    return f"""
        case
          when try_cast({alias}.account as integer) between {salary.first} and {salary.last}
            then 'Lønnskostnader'
          when try_cast({alias}.account as integer) between {depreciation.first} and {depreciation.last}
            then 'Avskrivninger'
          when try_cast({alias}.account as integer) between {adk.first} and {adk.last}
            then 'ADK'
          else null
        end
    """


def _budget_financing_case(alias: str) -> str:
    branches = "\n".join(
        f"when trim({alias}.dim_1) = '{section}' then '{financing}'"
        for section, financing in RULES.budget_financing_by_section.items()
    )
    return f"""
        case
          {branches}
          else '{RULES.budget_financing_default}'
        end
    """


def _actual_financing(alias: str) -> str:
    members = ", ".join(f"'{value}'" for value in RULES.combined_financing_members)
    return f"""
        case
          when trim({alias}.dim_4) in ({members}) then '{RULES.combined_financing_label}'
          else coalesce(nullif(trim({alias}.dim_4), ''), 'Uten finansiering')
        end
    """


def _available_periods(conn: duckdb.DuckDBPyConnection, ledger_path: Path) -> list[str]:
    salary = RULES.account_categories["Lønnskostnader"]
    rows = conn.execute(
        f"""
        select periode
        from (
          select
            trim(period) as periode,
            max(try_cast(trans_date as date)) as siste_transaksjonsdato,
            count(*) filter (
              where try_cast(account as integer) between {salary.first} and {salary.last}
            ) as lonnsrader
          from read_parquet('{ledger_path.as_posix()}')
          where regexp_matches(trim(period), '^20[0-9]{{2}}(0[1-9]|1[0-2])$')
          group by trim(period)
        )
        where lonnsrader > 0
          and siste_transaksjonsdato >= last_day(strptime(periode || '01', '%Y%m%d'))
        order by periode
        """
    ).fetchall()
    periods = [str(row[0]) for row in rows]
    if not periods:
        raise ValueError("Fant ingen avsluttet periode i hovedboken")
    return periods


def _latest_period(conn: duckdb.DuckDBPyConnection, ledger_path: Path) -> str:
    return _available_periods(conn, ledger_path)[-1]


def _summary_frame(
    conn: duckdb.DuckDBPyConnection,
    ledger_path: Path,
    budget_header_path: Path,
    budget_value_path: Path,
    period: str,
    cash_ledger_path: Path | None = None,
) -> pd.DataFrame:
    previous_period = f"{int(period) - 1:06d}"
    year_start = f"{period[:4]}01"
    budget_version = f"{period[:4]}B"
    budget_amount = _budget_amount_sql(conn, budget_value_path)
    categories = pd.DataFrame(
        [
            {"kategori": "Lønnskostnader", "sortering": 1},
            {"kategori": "Avskrivninger", "sortering": 2},
            {"kategori": "ADK", "sortering": 3},
            {"kategori": "Driftskostnader", "sortering": 4},
        ]
    )

    actual = conn.execute(
        f"""
        with base as (
          select
            trim(dim_1) as seksjon,
            {_actual_financing('a')} as finansiering,
            {_category_case('a')} as kategori,
            trim(period) as period,
            try_cast(amount as double) as amount
          from read_parquet('{ledger_path.as_posix()}') a
          where trim(period) between ? and ?
        )
        select seksjon, finansiering, kategori, period, sum(amount) as amount_nok
        from base
        where kategori is not null
        group by seksjon, finansiering, kategori, period
        """,
        [year_start, period],
    ).df()
    budget = conn.execute(
        f"""
        select
          trim(h.dim_1) as seksjon,
          {_budget_financing_case('h')} as finansiering,
          {_category_case('h')} as kategori,
          trim(v.period) as period,
          sum({budget_amount}) as amount_nok
        from read_parquet('{budget_header_path.as_posix()}') h
        join read_parquet('{budget_value_path.as_posix()}') v using (trans_id)
        where h.version = '{budget_version}'
          and trim(v.period) between ? and ?
        group by h.dim_1, finansiering, kategori, v.period
        having kategori is not null
        """,
        [year_start, period],
    ).df()

    report_sections = sorted(
        {
            str(value).strip()
            for value in pd.concat([actual["seksjon"], budget["seksjon"]], ignore_index=True)
            .dropna()
            .tolist()
            if str(value).strip().isdigit()
            and len(str(value).strip()) == 3
            and str(value).strip() != "999"
        }
        | set(SECTIONS)
    )

    def aggregate_scope(frame: pd.DataFrame, value_name: str) -> pd.DataFrame:
        section_rows = frame[frame["seksjon"].isin(report_sections)].copy()
        section_rows["omfang"] = "Seksjon"
        section_rows["omfang_id"] = section_rows["seksjon"]
        nkom_rows = frame.copy()
        nkom_rows["omfang"] = "Nkom"
        nkom_rows["omfang_id"] = "Nkom"
        combined = pd.concat([section_rows, nkom_rows], ignore_index=True)
        return (
            combined.groupby(
                ["omfang", "omfang_id", "finansiering", "kategori", "period"],
                as_index=False,
            )["amount_nok"]
            .sum()
            .rename(columns={"amount_nok": value_name})
        )

    actual_scoped = aggregate_scope(actual, "hovedbok_nok")
    budget_scoped = aggregate_scope(budget, "budsjett_nok")
    keys = ["omfang", "omfang_id", "finansiering", "kategori", "period"]
    monthly = actual_scoped.merge(budget_scoped, on=keys, how="outer")

    dimension_rows = monthly[["omfang", "omfang_id", "finansiering"]].drop_duplicates()
    dimension_rows = dimension_rows.merge(categories, how="cross")

    def period_values(source: pd.DataFrame, target_period: str, suffix: str) -> pd.DataFrame:
        subset = source[source["period"] == target_period].copy()
        return subset.drop(columns="period").rename(
            columns={
                "hovedbok_nok": f"hovedbok_{suffix}_nok",
                "budsjett_nok": f"budsjett_{suffix}_nok",
            }
        )

    current = period_values(monthly, period, "maaned")
    previous = period_values(monthly, previous_period, "forrige")
    ytd = (
        monthly.groupby(
            ["omfang", "omfang_id", "finansiering", "kategori"], as_index=False
        )[["hovedbok_nok", "budsjett_nok"]]
        .sum(min_count=1)
        .rename(
            columns={
                "hovedbok_nok": "hovedbok_hittil_nok",
                "budsjett_nok": "budsjett_hittil_nok",
            }
        )
    )
    join_keys = ["omfang", "omfang_id", "finansiering", "kategori"]
    result = dimension_rows.merge(current, on=join_keys, how="left")
    result = result.merge(previous, on=join_keys, how="left")
    result = result.merge(ytd, on=join_keys, how="left")

    operating_costs = result[
        result["kategori"].isin(["Lønnskostnader", "Avskrivninger", "ADK"])
    ].copy()
    total = (
        operating_costs.groupby(
            ["omfang", "omfang_id", "finansiering"], as_index=False
        )[
            [
                "hovedbok_maaned_nok",
                "budsjett_maaned_nok",
                "hovedbok_forrige_nok",
                "budsjett_forrige_nok",
                "hovedbok_hittil_nok",
                "budsjett_hittil_nok",
            ]
        ]
        .sum(min_count=1)
    )
    total["kategori"] = "Driftskostnader"
    total["sortering"] = 4
    result = pd.concat(
        [result[result["kategori"] != "Driftskostnader"], total], ignore_index=True
    )

    # At an existing ledger/budget grain, no matching posting means zero. It is
    # not a missing source. Genuine source gaps are represented separately.
    measure_columns = [
        "hovedbok_maaned_nok",
        "budsjett_maaned_nok",
        "hovedbok_forrige_nok",
        "budsjett_forrige_nok",
        "hovedbok_hittil_nok",
        "budsjett_hittil_nok",
    ]
    result[measure_columns] = result[measure_columns].fillna(0.0)
    for span in ("maaned", "forrige", "hittil"):
        result[f"avvik_{span}_nok"] = (
            result[f"budsjett_{span}_nok"] - result[f"hovedbok_{span}_nok"]
        )
    result["periode"] = period
    result["forrige_periode"] = previous_period
    result["budsjettversjon"] = budget_version
    result["kildestatus"] = "Beregnet"

    cash_source = cash_ledger_path or ledger_path
    cash_period_field = "pay_period" if cash_ledger_path else "period"
    cash_amount_field = "cash_amount" if cash_ledger_path else "amount"
    cash_712 = conn.execute(
        f"""
        select
          {_actual_financing('a')} as finansiering,
          coalesce(sum(try_cast({cash_amount_field} as double)) filter (where trim({cash_period_field}) = ?), 0) as maaned,
          coalesce(sum(try_cast({cash_amount_field} as double)) filter (where trim({cash_period_field}) = ?), 0) as forrige,
          coalesce(sum(try_cast({cash_amount_field} as double)), 0) as hittil
        from read_parquet('{cash_source.as_posix()}') a
        where trim(dim_1) = '{RULES.cash.section}'
          and trim(account) = '{RULES.cash.account}'
          and trim(dim_4) = '{RULES.cash.financing}'
          and trim({cash_period_field}) between ? and ?
        group by finansiering
        """,
        [period, previous_period, year_start, period],
    ).df()
    cash_rows = []
    for cash in cash_712.itertuples(index=False):
        for scope, scope_id in (("Seksjon", RULES.cash.section), ("Nkom", "Nkom")):
            for category, sorting, current, previous, ytd_value in (
                ("Lønnskostnader", 1, 0.0, 0.0, 0.0),
                ("Tilskudd", 2, cash.maaned, cash.forrige, cash.hittil),
                ("Driftskostnader", 4, cash.maaned, cash.forrige, cash.hittil),
            ):
                cash_rows.append(
                    {
                        "omfang": scope,
                        "omfang_id": scope_id,
                        "finansiering": cash.finansiering,
                        "kategori": category,
                        "sortering": sorting,
                        "hovedbok_maaned_nok": current,
                        "budsjett_maaned_nok": RULES.cash.budget_nok,
                        "avvik_maaned_nok": RULES.cash.budget_nok - current,
                        "hovedbok_forrige_nok": previous,
                        "budsjett_forrige_nok": RULES.cash.budget_nok,
                        "avvik_forrige_nok": RULES.cash.budget_nok - previous,
                        "hovedbok_hittil_nok": ytd_value,
                        "budsjett_hittil_nok": RULES.cash.budget_nok,
                        "avvik_hittil_nok": RULES.cash.budget_nok - ytd_value,
                        "periode": period,
                        "forrige_periode": previous_period,
                        "budsjettversjon": budget_version,
                        "kildestatus": (
                            f"Kontantgrunnlag: {cash_source.name}.{cash_amount_field}, "
                            f"periodisert med {cash_period_field}"
                        ),
                    }
                )
    if cash_rows:
        result = pd.concat([result, pd.DataFrame(cash_rows)], ignore_index=True)
    existing_sections = set(
        result.loc[result["omfang"] == "Seksjon", "omfang_id"].astype(str)
    )
    placeholder_rows = []
    for section in sorted(set(SECTIONS) - existing_sections):
        for category, sorting in (
            ("Lønnskostnader", 1),
            ("Avskrivninger", 2),
            ("ADK", 3),
            ("Driftskostnader", 4),
        ):
            placeholder_rows.append(
                {
                    "omfang": "Seksjon",
                    "omfang_id": section,
                    "finansiering": "Ikke tilgjengelig",
                    "kategori": category,
                    "sortering": sorting,
                    "periode": period,
                    "forrige_periode": previous_period,
                    "budsjettversjon": budget_version,
                    "kildestatus": "Kontantkilde per seksjon mangler",
                }
            )
    if placeholder_rows:
        result = pd.concat([result, pd.DataFrame(placeholder_rows)], ignore_index=True)
    return result.sort_values(
        ["omfang", "omfang_id", "finansiering", "sortering"], na_position="last"
    ).reset_index(drop=True)


def _invoice_frame(
    conn: duckdb.DuckDBPyConnection,
    workflow_path: Path,
    ledger_path: Path,
) -> tuple[pd.DataFrame, pd.Timestamp | None]:
    active_status = RULES.workflow_candidates.active_status
    completed_actions = ", ".join(
        f"'{action}'" for action in RULES.workflow_candidates.completed_actions
    )
    snapshot = conn.execute(
        f"""
        select max(coalesce(
          try_cast(action_date as timestamp),
          try_cast(ready_date as timestamp),
          try_cast(distr_date as timestamp)
        ))
        from read_parquet('{workflow_path.as_posix()}')
        """
    ).fetchone()[0]
    invoices = conn.execute(
        f"""
        with workflow_base as (
          select
            trim(col2_value) as fakturanr,
            oid,
            action_code,
            wf_status,
            trim(col1_value) as leverandor_navn,
            try_cast(col5_value as double) as belop_nok,
            trim(col6_value) as vist_konto,
            logged_values,
            try_cast(action_date as timestamp) as action_ts,
            coalesce(
              try_cast(action_date as timestamp),
              try_cast(ready_date as timestamp),
              try_cast(distr_date as timestamp)
            ) as event_ts
          from read_parquet('{workflow_path.as_posix()}')
          where col2_descr = 'Fakturanr'
            and trim(coalesce(col2_value, '')) <> ''
        ),
        per_flow as (
          select
            fakturanr,
            oid,
            arg_max(action_code, action_ts) as siste_handling,
            max(action_ts) as siste_handling_tid,
            arg_max(logged_values, action_ts) filter (where logged_values is not null) as dimensjoner,
            arg_max(leverandor_navn, event_ts) as leverandor_navn,
            arg_max(belop_nok, action_ts) filter (where logged_values is not null) as belop_nok,
            arg_max(vist_konto, action_ts) filter (where logged_values is not null) as vist_konto,
            count(*) filter (where wf_status = '{active_status}') as aktive_oppgaver
          from workflow_base
          group by fakturanr, oid
        ),
        ledger as (
          select distinct trim(ext_inv_ref) as fakturanr
          from read_parquet('{ledger_path.as_posix()}')
          where trim(coalesce(ext_inv_ref, '')) <> ''
        )
        select
          p.fakturanr,
          p.oid,
          p.leverandor_navn,
          p.belop_nok,
          coalesce(nullif(regexp_extract(p.dimensjoner, 'A0=([^;]+)', 1), ''), p.vist_konto) as konto,
          nullif(regexp_extract(p.dimensjoner, 'C1=([^;]+)', 1), '') as seksjon,
          nullif(regexp_extract(p.dimensjoner, 'B0=([^;]+)', 1), '') as prosjektnr,
          nullif(regexp_extract(p.dimensjoner, 'R00=([^;]+)', 1), '') as finansiering,
          case
            when date_diff('day', cast(p.siste_handling_tid as date), cast(? as date))
              between 0 and {RULES.workflow_candidates.stale_after_days}
              then 'Aktuell kandidat'
            else 'Historisk workflowpost'
          end as maanedsavslutningsstatus,
          case
            when date_diff('day', cast(p.siste_handling_tid as date), cast(? as date))
              between 0 and {RULES.workflow_candidates.stale_after_days}
              then 'Ikke bokført i snapshot; har ACT-oppgave; siste fullførte handling er '
                || p.siste_handling
            else 'Holdt utenfor arbeidslisten: siste registrerte handling er eldre enn '
                || {RULES.workflow_candidates.stale_after_days} || ' dager; ACT-statusen må bekreftes mot fakturasystemet'
          end as statusgrunnlag,
          p.siste_handling,
          p.siste_handling_tid,
          p.aktive_oppgaver,
          date_diff('day', cast(p.siste_handling_tid as date), cast(? as date)) as alder_dager,
          date_diff('day', cast(p.siste_handling_tid as date), cast(? as date))
            between 0 and {RULES.workflow_candidates.stale_after_days} as er_aktuell,
          'Ikke bokført i mottatt hovedbokssnapshot' as bokforingskontroll
        from per_flow p
        left join ledger l using (fakturanr)
        where l.fakturanr is null
          and p.aktive_oppgaver > 0
          and p.siste_handling in ({completed_actions})
        order by p.siste_handling_tid desc, p.fakturanr
        """,
        [snapshot, snapshot, snapshot, snapshot],
    ).df()
    return invoices, pd.Timestamp(snapshot) if snapshot else None


def _lookup(
    summary: pd.DataFrame,
    scope: str,
    financing: str,
    category: str,
    column: str,
) -> float | None:
    rows = summary[
        (summary["omfang_id"] == scope)
        & (summary["finansiering"] == financing)
        & (summary["kategori"] == category)
    ]
    if rows.empty or pd.isna(rows.iloc[0][column]):
        return None
    return float(rows.iloc[0][column])


def _excel_value(value: object, missing_label: str = "–") -> object:
    """Keep missing source values visible instead of making the workbook look empty."""
    if value is None or pd.isna(value):
        return missing_label
    return float(value)


def _write_excel_value(ws, coordinate: str, value: object) -> None:
    ws[coordinate] = _excel_value(value)
    if isinstance(ws[coordinate].value, (int, float)):
        # The template's accounting format renders zero as a dash. Explicit zero
        # is clearer here because it distinguishes no postings from no source.
        ws[coordinate].number_format = '#,##0;[Red]-#,##0;0'


def _sum_if_complete(*values: float | None) -> float | None:
    if any(value is None for value in values):
        return None
    return float(sum(values))


def _fill_summary_block(
    ws,
    summary: pd.DataFrame,
    scope: str,
    financing: str,
    current_columns: tuple[str, str, str],
    ytd_columns: tuple[str, str, str] | None,
    previous_columns: tuple[str, str, str] | None,
    rows: dict[str, int],
) -> None:
    for category, row_number in rows.items():
        for column_letter, field in zip(
            current_columns,
            ("hovedbok_maaned_nok", "budsjett_maaned_nok", "avvik_maaned_nok"),
        ):
            _write_excel_value(
                ws,
                f"{column_letter}{row_number}",
                _lookup(summary, scope, financing, category, field),
            )
        if ytd_columns:
            for column_letter, field in zip(
                ytd_columns,
                ("hovedbok_hittil_nok", "budsjett_hittil_nok", "avvik_hittil_nok"),
            ):
                _write_excel_value(
                    ws,
                    f"{column_letter}{row_number}",
                    _lookup(summary, scope, financing, category, field),
                )
        if previous_columns:
            for column_letter, field in zip(
                previous_columns,
                ("hovedbok_forrige_nok", "budsjett_forrige_nok", "avvik_forrige_nok"),
            ):
                _write_excel_value(
                    ws,
                    f"{column_letter}{row_number}",
                    _lookup(summary, scope, financing, category, field),
                )


def _fill_invoice_rows(ws, invoices: pd.DataFrame, section: str) -> int:
    config = {
        "711": (range(14, 18), ("B", "C", "D", "E", "F")),
        "721": (range(14, 18), ("B", "C", "D", "E", "F")),
        "731": (range(16, 20), ("B", "C", "D", "E", "F")),
        "741": (range(19, 23), ("B", "C", "D", "E", "F", "G", "H")),
    }
    if section not in config:
        return 0
    row_range, columns = config[section]
    for row_number in row_range:
        ws.row_dimensions[row_number].hidden = False
        for column in columns:
            ws[f"{column}{row_number}"] = None

    relevant = invoices[invoices["seksjon"] == section].head(len(row_range))
    if relevant.empty:
        first_row = row_range.start
        ws[f"B{first_row}"] = "Ingen aktuelle fakturaer i mottatt workflow-snapshot"
        for row_number in list(row_range)[1:]:
            ws.row_dimensions[row_number].hidden = True
    else:
        for row_number in list(row_range)[len(relevant) :]:
            ws.row_dimensions[row_number].hidden = True
    for row_number, row in zip(row_range, relevant.itertuples(index=False)):
        if section == "741":
            values = [
                row.konto,
                row.seksjon,
                row.prosjektnr,
                row.finansiering,
                row.leverandor_navn,
                row.belop_nok,
                f"{row.fakturanr} – {row.maanedsavslutningsstatus}",
            ]
        else:
            values = [
                row.konto,
                row.seksjon,
                row.finansiering,
                row.belop_nok,
                f"{row.leverandor_navn} – {row.fakturanr} – {row.maanedsavslutningsstatus}",
            ]
        for column, value in zip(columns, values):
            ws[f"{column}{row_number}"] = value
    total_cell = {"711": "E18", "721": "E18", "731": "E20", "741": "G23"}[section]
    amount_total = relevant["belop_nok"].sum(min_count=1)
    ws[total_cell] = 0.0 if pd.isna(amount_total) else float(amount_total)
    return len(relevant)


def _add_nkom_sheet(wb, summary: pd.DataFrame, period: str) -> None:
    if "Nkom per finansiering" in wb.sheetnames:
        del wb["Nkom per finansiering"]
    ws = wb.create_sheet("Nkom per finansiering")
    ws.append(
        [
            "Periode",
            "Finansiering",
            "Kategori",
            "Hovedbok måned",
            "Budsjett måned",
            "Diff måned",
            "Hovedbok hittil i år",
            "Budsjett hittil i år",
            "Diff hittil i år",
            "Budsjettversjon",
        ]
    )
    value_columns = [
        "hovedbok_maaned_nok",
        "budsjett_maaned_nok",
        "hovedbok_hittil_nok",
        "budsjett_hittil_nok",
    ]
    rows = summary[
        (summary["omfang"] == "Nkom")
        & (summary[value_columns].fillna(0).abs().sum(axis=1) > 0.005)
    ]
    for row in rows.itertuples(index=False):
        ws.append(
            [
                period,
                row.finansiering,
                row.kategori,
                _excel_value(row.hovedbok_maaned_nok),
                _excel_value(row.budsjett_maaned_nok),
                _excel_value(row.avvik_maaned_nok),
                _excel_value(row.hovedbok_hittil_nok),
                _excel_value(row.budsjett_hittil_nok),
                _excel_value(row.avvik_hittil_nok),
                RULES.budget_version,
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(4, 10):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for value_cell in cell:
                value_cell.number_format = '#,##0;[Red]-#,##0;0'
    for column, width in {"A": 12, "B": 20, "C": 20, "D": 20, "E": 20, "F": 18, "G": 23, "H": 23, "I": 18, "J": 18}.items():
        ws.column_dimensions[column].width = width


def _add_section_financing_sheet(wb, summary: pd.DataFrame, period: str) -> None:
    if "Seksjon per finansiering" in wb.sheetnames:
        del wb["Seksjon per finansiering"]
    ws = wb.create_sheet("Seksjon per finansiering")
    ws.append(
        [
            "Periode",
            "Seksjon",
            "Finansiering",
            "Kategori",
            "Hovedbok måned",
            "Budsjett måned",
            "Diff måned",
            "Hovedbok forrige måned",
            "Budsjett forrige måned",
            "Diff forrige måned",
            "Hovedbok hittil i år",
            "Budsjett hittil i år",
            "Diff hittil i år",
            "Budsjettversjon",
            "Kildestatus",
        ]
    )
    value_columns = [
        "hovedbok_maaned_nok",
        "budsjett_maaned_nok",
        "hovedbok_forrige_nok",
        "budsjett_forrige_nok",
        "hovedbok_hittil_nok",
        "budsjett_hittil_nok",
    ]
    rows = summary[
        (summary["omfang"] == "Seksjon")
        & (summary[value_columns].fillna(0).abs().sum(axis=1) > 0.005)
    ]
    for row in rows.itertuples(index=False):
        ws.append(
            [
                period,
                row.omfang_id,
                row.finansiering,
                row.kategori,
                _excel_value(row.hovedbok_maaned_nok),
                _excel_value(row.budsjett_maaned_nok),
                _excel_value(row.avvik_maaned_nok),
                _excel_value(row.hovedbok_forrige_nok),
                _excel_value(row.budsjett_forrige_nok),
                _excel_value(row.avvik_forrige_nok),
                _excel_value(row.hovedbok_hittil_nok),
                _excel_value(row.budsjett_hittil_nok),
                _excel_value(row.avvik_hittil_nok),
                RULES.budget_version,
                row.kildestatus,
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(5, 14):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for value_cell in cell:
                value_cell.number_format = '#,##0;[Red]-#,##0;0'
    for column, width in {"A": 12, "B": 12, "C": 20, "D": 20, "E": 20, "F": 20, "G": 18, "H": 22, "I": 22, "J": 18, "K": 22, "L": 22, "M": 18, "N": 18, "O": 52}.items():
        ws.column_dimensions[column].width = width


def _add_draft_sheet(wb, invoices: pd.DataFrame, snapshot: pd.Timestamp | None) -> None:
    if "Bilagsutkast kontroll" in wb.sheetnames:
        del wb["Bilagsutkast kontroll"]
    ws = wb.create_sheet("Bilagsutkast kontroll")
    ws.append(["UTKAST – IKKE BOKFØR UTEN FAGLIG KONTROLL"])
    ws.append(
        [
            "Fakturanummer",
            "Leverandør",
            "Status",
            "Statusgrunnlag",
            "Konto",
            "Seksjon",
            "Prosjekt",
            "Finansiering",
            "Beløp",
            "Siste handling",
            "Alder dager",
            "Aktualitet",
            "Bokføringskontroll",
        ]
    )
    for row in invoices.itertuples(index=False):
        ws.append(
            [
                row.fakturanr,
                row.leverandor_navn,
                row.maanedsavslutningsstatus,
                row.statusgrunnlag,
                row.konto,
                row.seksjon,
                row.prosjektnr,
                row.finansiering,
                row.belop_nok,
                row.siste_handling_tid,
                row.alder_dager,
                row.maanedsavslutningsstatus,
                row.bokforingskontroll,
            ]
        )
    ws["A1"].font = Font(bold=True, color="9C0006")
    ws["M1"] = "Workflow-snapshot"
    ws["N1"] = None if snapshot is None else snapshot.to_pydatetime()
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:M{max(2, ws.max_row)}"


def _add_cash_712_detail_sheet(wb, root: Path, period: str) -> None:
    sheet_name = f"{RULES.cash.section} kontantdetaljer"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    sources = task3_sources()
    ledger_path = sources.cash_ledger or sources.ledger
    period_field = "pay_period" if sources.cash_ledger else "period"
    amount_field = "cash_amount" if sources.cash_ledger else "amount"
    conn = duckdb.connect()
    try:
        rows = conn.execute(
            f"""
            select
              trim({period_field}) as periode,
              trim(account) as konto,
              trim(dim_1) as seksjon,
              trim(dim_2) as prosjekt,
              trim(dim_4) as finansiering,
              description as beskrivelse,
              try_cast({amount_field} as double) as belop_nok
            from read_parquet('{ledger_path.as_posix()}')
            where trim(dim_1) = '{RULES.cash.section}'
              and trim(account) = '{RULES.cash.account}'
              and trim(dim_4) = '{RULES.cash.financing}'
              and trim({period_field}) between ? and ?
            order by periode, prosjekt, belop_nok desc
            """,
            [f"{period[:4]}01", period],
        ).df()
    finally:
        conn.close()

    ws = wb.create_sheet(sheet_name)
    ws.append(
        [
            "Periode",
            "Konto",
            "Seksjon",
            "Prosjekt",
            "Finansiering",
            "Beskrivelse",
            "Beløp NOK",
        ]
    )
    for row in rows.itertuples(index=False):
        ws.append(list(row))
    ws.append([None, None, None, None, None, "Sum hittil i år", float(rows["belop_nok"].sum())])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(1, ws.max_row - 1)}"
    ws.column_dimensions["A"].width = 12
    for column in ("B", "C", "D", "E"):
        ws.column_dimensions[column].width = 14
    ws.column_dimensions["F"].width = 88
    ws.column_dimensions["G"].width = 20
    for cell in ws["G"][1:]:
        cell.number_format = '#,##0;[Red]-#,##0;0'


def _fill_workbook(
    root: Path,
    summary: pd.DataFrame,
    invoices: pd.DataFrame,
    period: str,
    snapshot: pd.Timestamp | None,
) -> Path:
    sources = task3_sources()
    template = sources.monthly_close_template
    if not template.exists():
        raise FileNotFoundError(f"Mangler månedsavslutningsmal: {template}")
    output_dir = sources.generated_dir / "reports"
    static_dir = root / "static"
    output_dir.mkdir(parents=True, exist_ok=True)
    static_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"manedsavslutning_{period[:4]}-{period[4:]}.xlsx"

    wb = load_workbook(template)
    current_invoices = invoices[invoices["er_aktuell"].fillna(False)].copy()
    current_name = MONTH_NAMES[int(period[4:])]
    previous_name = MONTH_NAMES[int(period[4:]) - 1] if int(period[4:]) > 1 else "Desember"

    section_config = {
        "711": ("711 - SID", ("C", "D", "E"), ("J", "K", "L"), ("N", "O", "P"), {"Lønnskostnader": 7, "ADK": 8, "Driftskostnader": 9}),
        "721": ("721 - SA", ("C", "D", "E"), ("I", "J", "K"), ("M", "N", "O"), {"Lønnskostnader": 7, "ADK": 8, "Driftskostnader": 9}),
        "731": ("731 - SB", ("C", "D", "E"), ("I", "J", "K"), ("M", "N", "O"), {"Lønnskostnader": 7, "ADK": 8, "Driftskostnader": 9}),
        "741": ("741 - SC", ("C", "D", "E"), ("G", "H", "I"), ("L", "M", "N"), {"Lønnskostnader": 7, "ADK": 8, "Driftskostnader": 9}),
    }
    for section, (sheet_name, current_cols, ytd_cols, previous_cols, rows) in section_config.items():
        ws = wb[sheet_name]
        ws["C3"] = current_name
        for cell in ("C5", "D5"):
            ws[cell] = current_name
        previous_label_columns = previous_cols[:2]
        for column in previous_label_columns:
            ws[f"{column}5"] = previous_name
        _fill_summary_block(
            ws,
            summary,
            section,
            "154301",
            current_cols,
            ytd_cols,
            previous_cols,
            rows,
        )
        _fill_invoice_rows(ws, current_invoices, section)

    # Remove dated example comments from the supplied template. They must not be
    # mistaken for current-month facts.
    wb["711 - SID"]["B20"] = "Tilleggsforslag kan ikke oppdateres før prosjektmapping er godkjent"
    wb["721 - SA"]["B19"] = "Tidligere eksempelkommentarer er fjernet – ingen aktuell kommentar i kilden"
    for row_number in (20, 21):
        wb["721 - SA"].row_dimensions[row_number].hidden = True
    wb["731 - SB"]["B22"] = "Tilleggsforslag kan ikke oppdateres før prosjektmapping er godkjent"
    wb["731 - SB"]["B23"] = None
    wb["731 - SB"].row_dimensions[23].hidden = True

    total_ws = wb["Totalt eks 712"]
    total_ws["C3"] = current_name
    for cell in ("C5", "D5", "E5"):
        total_ws[cell] = current_name
    total_rows = {"Lønnskostnader": 7, "ADK": 8, "Driftskostnader": 9}
    section_total = summary[
        (summary["omfang"] == "Seksjon")
        & (summary["omfang_id"].isin(["711", "721", "731", "741"]))
        & (summary["finansiering"] == "154301")
    ]
    synthetic = (
        section_total.groupby(["finansiering", "kategori", "sortering"], as_index=False)[
            [column for column in summary.columns if column.endswith("_nok")]
        ]
        .sum(min_count=1)
    )
    synthetic["omfang"] = "Seksjon"
    synthetic["omfang_id"] = f"Totalt eks {RULES.cash.section}"
    _fill_summary_block(
        total_ws,
        synthetic,
        f"Totalt eks {RULES.cash.section}",
        RULES.budget_financing_default,
        ("C", "D", "E"),
        ("G", "H", "I"),
        None,
        total_rows,
    )
    total_ws["C14"] = current_name

    cash_ws = wb[RULES.cash.section]
    cash_ws["B3"] = f"Resultat {current_name} – post 70, konto {RULES.cash.account}"
    cash_ws["B6"] = f"Tilskudd (konto {RULES.cash.account})"
    cash_ws["F5"] = "Hovedbok h.i.å."
    cash_ws["G5"] = "Budsjett h.i.å."
    cash_ws["H5"] = "Diff h.i.å."
    cash_ws["J5"] = f"Hovedbok {previous_name}"
    cash_ws["K5"] = f"Budsjett {previous_name}"
    cash_ws["L5"] = f"Diff {previous_name}"
    cash_financing = RULES.cash.financing
    cash_categories = {"Tilskudd": 6}
    _fill_summary_block(
        cash_ws,
        summary,
        RULES.cash.section,
        cash_financing,
        ("C", "D", "E"),
        ("F", "G", "H"),
        ("J", "K", "L"),
        cash_categories,
    )
    for row_number in (7, 8):
        cash_ws.row_dimensions[row_number].hidden = True
    cash_ws["B10"] = (
        f"Foreløpig kontantgrunnlag: hovedbok konto {RULES.cash.account} / "
        f"finansiering {RULES.cash.financing} – må faglig bekreftes"
    )
    cash_ws["B15"] = "Ingen avsetningsbilag registrert for perioden"
    for row_number in (17, 18, 19):
        cash_ws.row_dimensions[row_number].hidden = True

    section_741 = wb["741 - SC"]
    section_741["I19"] = None
    section_741["O7"] = None
    section_741["B4"] = "Post 154301"
    section_741["C12"] = current_name
    section_741["D12"] = current_name
    section_741_financings = summary.loc[
        (summary["omfang"] == "Seksjon")
        & (summary["omfang_id"] == "741")
        & (summary["finansiering"] != "154301")
        & summary[
            [
                "hovedbok_maaned_nok",
                "budsjett_maaned_nok",
                "hovedbok_hittil_nok",
                "budsjett_hittil_nok",
            ]
        ].notna().any(axis=1),
        "finansiering",
    ].drop_duplicates()
    secondary_financing = (
        str(section_741_financings.iloc[0]) if not section_741_financings.empty else None
    )
    if secondary_financing:
        section_741["B11"] = f"Post {secondary_financing}"
        section_741["B46"] = f"Post {secondary_financing}"
        _fill_summary_block(
            section_741,
            summary,
            "741",
            secondary_financing,
            ("C", "D", "E"),
            ("G", "H", "I"),
            None,
            {"ADK": 14, "Driftskostnader": 15},
        )
    else:
        section_741["B11"] = "Ingen øvrig finansiering i mottatte data"
        for row_number in (12, 13, 14, 15, 46, 47, 48, 50, 51):
            section_741.row_dimensions[row_number].hidden = True
    section_741["C38"] = current_name
    section_741["C47"] = current_name

    # The template reserves areas for future proposal classifications. The
    # received sources contain project numbers, but no approved mapping to these
    # three categories. Show that limitation once per area and collapse empty rows.
    for status_row, hidden_rows in (
        (25, (26, 27)),
        (29, (30, 31)),
        (33, (34, 35)),
    ):
        section_741[f"B{status_row}"] = "Ikke tilgjengelig – godkjent prosjektmapping mangler"
        for column in ("C", "D", "E", "F"):
            section_741[f"{column}{status_row}"] = "–"
        for row_number in hidden_rows:
            section_741.row_dimensions[row_number].hidden = True

    # Store post-accrual totals as values. Formula-only cells have no cached result
    # when openpyxl writes them and therefore look empty in browser/Excel previews.
    post_accrual = {}
    for section, (sheet_name, _, _, _, _) in section_config.items():
        ws = wb[sheet_name]
        salary = _lookup(summary, section, "154301", "Lønnskostnader", "hovedbok_maaned_nok")
        depreciation = _lookup(
            summary, section, "154301", "Avskrivninger", "hovedbok_maaned_nok"
        )
        adk = _lookup(summary, section, "154301", "ADK", "hovedbok_maaned_nok")
        invoice_total = current_invoices.loc[
            current_invoices["seksjon"] == section, "belop_nok"
        ].sum(min_count=1)
        invoice_total = 0.0 if pd.isna(invoice_total) else float(invoice_total)
        if not RULES.workflow_candidates.include_amounts_in_calculations:
            invoice_total = 0.0
        adjusted_adk = None if adk is None else adk + invoice_total
        adjusted_total = _sum_if_complete(salary, depreciation, adjusted_adk)
        post_accrual[section] = (salary, adjusted_adk, adjusted_total)
        target_rows = (41, 42, 43) if section == "741" else (
            (29, 30, 31) if section == "731" else (26, 27, 28)
        )
        for row_number, value in zip(target_rows, post_accrual[section]):
            _write_excel_value(ws, f"C{row_number}", value)

    total_salary = _sum_if_complete(*(post_accrual[s][0] for s in section_config))
    total_adk = _sum_if_complete(*(post_accrual[s][1] for s in section_config))
    total_after = _sum_if_complete(*(post_accrual[s][2] for s in section_config))
    for row_number, value in zip((17, 18, 19), (total_salary, total_adk, total_after)):
        _write_excel_value(total_ws, f"C{row_number}", value)

    secondary_adk = (
        _lookup(summary, "741", secondary_financing, "ADK", "hovedbok_maaned_nok")
        if secondary_financing
        else None
    )
    secondary_total = (
        _lookup(
            summary,
            "741",
            secondary_financing,
            "Driftskostnader",
            "hovedbok_maaned_nok",
        )
        if secondary_financing
        else None
    )
    _write_excel_value(section_741, "C50", secondary_adk)
    _write_excel_value(section_741, "C51", secondary_total)

    voucher_ws = wb["Avsetningsbilag"]
    voucher_ws["A74"] = "IKKE BOKFØRINGSKLAR – bruk fanen 'Bilagsutkast kontroll' og gjennomfør faglig kontroll"
    voucher_ws["A74"].font = Font(bold=True, color="9C0006")
    voucher_ws.sheet_state = "hidden"

    _add_nkom_sheet(wb, summary, period)
    _add_section_financing_sheet(wb, summary, period)
    _add_cash_712_detail_sheet(wb, root, period)
    _add_draft_sheet(wb, invoices, snapshot)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output)
    shutil.copy2(output, static_dir / "manedsavslutning-siste.xlsx")
    shutil.copy2(output, static_dir / output.name)
    return output


def build_monthly_close(root: Path) -> MonthlyCloseResult:
    sources = task3_sources()
    ledger_path = sources.ledger
    workflow_path = sources.workflow
    budget_header_path = sources.budget_header
    budget_value_path = sources.budget_values
    for path in (ledger_path, workflow_path, budget_header_path, budget_value_path):
        if not path.exists():
            raise FileNotFoundError(f"Mangler operativ kilde: {path}")

    conn = duckdb.connect()
    try:
        periods = _available_periods(conn, ledger_path)
        period = periods[-1]
        summaries = [
            _summary_frame(
                conn,
                ledger_path,
                budget_header_path,
                budget_value_path,
                candidate,
                sources.cash_ledger,
            )
            for candidate in periods
        ]
        summary = pd.concat(summaries, ignore_index=True)
        latest_summary = summary[summary["periode"] == period].copy()
        invoices, snapshot = _invoice_frame(conn, workflow_path, ledger_path)
    finally:
        conn.close()

    workbook_path = _fill_workbook(root, latest_summary, invoices, period, snapshot)
    current_invoices = invoices[invoices["er_aktuell"].fillna(False)]
    missing_dimensions = int(
        current_invoices[["konto", "seksjon", "finansiering"]]
        .isna()
        .any(axis=1)
        .sum()
    )
    stale_days = RULES.workflow_candidates.stale_after_days
    stale_invoices = int((~invoices["er_aktuell"].fillna(False)).sum())
    actions_text = " eller ".join(RULES.workflow_candidates.completed_actions)
    validations = pd.DataFrame(
        [
            {"kontroll": "Aktuell periode", "status": "ok", "antall": 1, "detalj": f"Siste hovedboksperiode er {period}."},
            {"kontroll": "Seksjoner i mal", "status": "ok", "antall": len(SECTIONS), "detalj": f"Malen dekker {', '.join(SECTIONS)}."},
            {"kontroll": "Budsjettversjon", "status": "ok", "antall": len({str(value) for value in summary['budsjettversjon']}), "detalj": f"Opprinnelig budsjett per rapportår er brukt ({summary['budsjettversjon'].min()}–{summary['budsjettversjon'].max()})."},
            {"kontroll": "Aktuelle kandidater til fakturakontroll", "status": "warning" if len(current_invoices) else "ok", "antall": len(current_invoices), "detalj": f"Ikke bokført i snapshot, har {RULES.workflow_candidates.active_status}-oppgave, siste fullførte handling {actions_text} og er høyst {stale_days} dager gammel. Endelig fakturastatus må godkjennes."},
            {"kontroll": "Manglende fakturadimensjoner", "status": "warning" if missing_dimensions else "ok", "antall": missing_dimensions, "detalj": "Konto, seksjon og finansiering leses fra workflow-loggen."},
            {"kontroll": "Historiske workflowposter", "status": "warning" if stale_invoices else "ok", "antall": stale_invoices, "detalj": f"Poster eldre enn {stale_days} dager holdes utenfor arbeidslisten og må bekreftes mot fakturasystemet."},
            {"kontroll": f"Kontantgrunnlag seksjon {RULES.cash.section}", "status": "ok" if sources.cash_ledger else "warning", "antall": len(latest_summary[(latest_summary["omfang"] == "Seksjon") & (latest_summary["omfang_id"] == RULES.cash.section)]), "detalj": f"Konto {RULES.cash.account} / finansiering {RULES.cash.financing} er periodisert med pay_period i acatrans." if sources.cash_ledger else f"Konto {RULES.cash.account} / finansiering {RULES.cash.financing} er brukt foreløpig og må bekreftes som riktig kontantgrunnlag."},
            {"kontroll": "Beløp fra workflow", "status": "warning", "antall": len(current_invoices), "detalj": "Workflowbeløp påvirker ikke beregnede regnskapstall. Regelen er ikke faglig godkjent."},
        ]
    )
    return MonthlyCloseResult(summary, invoices, validations, period, workbook_path)


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parents[1]
    result = build_monthly_close(project_root)
    print(f"Skrev {result.workbook_path}")
    current_count = int(result.invoices["er_aktuell"].fillna(False).sum())
    print(f"Periode: {result.period}; {current_count} aktuelle fakturaer til kontroll")
