from __future__ import annotations

from pathlib import Path

import duckdb
from openpyxl import load_workbook

try:
    from .project_data import task3_sources
    from .task3_rules import load_task3_rules
except ImportError:
    from project_data import task3_sources
    from task3_rules import load_task3_rules


ROOT = Path(__file__).resolve().parents[1]
RULES = load_task3_rules()
SOURCES = task3_sources()
GENERATED_DIR = SOURCES.generated_dir / "web"
STATUS_PATH = GENERATED_DIR / "workflow_invoice_status.parquet"
EVENTS_PATH = GENERATED_DIR / "workflow_invoice_events.parquet"
VALIDATION_PATH = GENERATED_DIR / "workflow_invoice_validation.parquet"
METADATA_PATH = GENERATED_DIR / "workflow_source_metadata.parquet"
MONTHLY_SUMMARY_PATH = GENERATED_DIR / "monthly_close_summary.parquet"
MONTHLY_INVOICES_PATH = GENERATED_DIR / "monthly_close_invoices.parquet"
MONTHLY_VALIDATION_PATH = GENERATED_DIR / "monthly_close_validation.parquet"
MONTHLY_WORKBOOK_PATH = ROOT / "static" / "manedsavslutning-siste.xlsx"
LEDGER_PATH = SOURCES.ledger
WORKFLOW_PATH = SOURCES.workflow


def main() -> None:
    required_paths = [
        STATUS_PATH,
        EVENTS_PATH,
        VALIDATION_PATH,
        METADATA_PATH,
        MONTHLY_SUMMARY_PATH,
        MONTHLY_INVOICES_PATH,
        MONTHLY_VALIDATION_PATH,
        MONTHLY_WORKBOOK_PATH,
    ]
    if any(not path.exists() for path in required_paths):
        raise SystemExit("Genererte workflowdata mangler. Kjør npm run prepare:data først.")

    conn = duckdb.connect()
    try:
        summary = conn.execute(
            f"""
            select
              count(*) as fakturaer,
              count(distinct fakturanr) as unike_fakturaer,
              count(*) filter (where fakturanr is null or trim(fakturanr) = '') as uten_fakturanr,
              count(*) filter (
                where koblingskvalitet not in ('Sikker', 'Mulig', 'Tvetydig', 'Ikke matchet')
              ) as ugyldig_koblingskvalitet,
              count(*) filter (where koblingskvalitet = 'Sikker') as sikre,
              count(*) filter (where koblingskvalitet = 'Tvetydig') as tvetydige,
              count(*) filter (where koblingskvalitet = 'Ikke matchet') as ikke_matchet
            from read_parquet('{STATUS_PATH.as_posix()}')
            """
        ).fetchone()
        safe_map_clause = (
            "or coalesce(mapped_regnskapsbilag, 0) > 0" if SOURCES.ledger_map else ""
        )
        invalid_safe = conn.execute(
            f"""
            select count(*)
            from read_parquet('{STATUS_PATH.as_posix()}')
            where koblingskvalitet = 'Sikker'
              and not (
                workflow_flyter = 1
                and workflow_leverandor_id_antall = 1
                and regnskap_leverandorer = 1
                and workflow_leverandor_id = regnskap_leverandor_id
                {safe_map_clause}
              )
            """
        ).fetchone()[0]
        validation_count = conn.execute(
            f"select count(*) from read_parquet('{VALIDATION_PATH.as_posix()}')"
        ).fetchone()[0]
        metadata_errors = conn.execute(
            f"""
            select count(*)
            from read_parquet('{METADATA_PATH.as_posix()}')
            where snapshot_status is null
               or workflow_fil_endret is null
               or regnskap_fil_endret is null
               or seneste_workflowhendelse is null
               or seneste_bilagsdato is null
               or seneste_hovedboksperiode is null
            """
        ).fetchone()[0]
        metadata_rows = conn.execute(
            f"select count(*) from read_parquet('{METADATA_PATH.as_posix()}')"
        ).fetchone()[0]
        event_summary = conn.execute(
            f"""
            select
              count(*) as hendelser,
              count(distinct fakturanr) as fakturaer,
              count(*) filter (where oid is null or trim(oid) = '') as uten_oid,
              count(*) filter (where hendelse_tid is null) as uten_hendelsestid,
              count(*) filter (
                where oppgavestatus_tekst is null or handling_tekst is null
              ) as uten_forklaring
            from read_parquet('{EVENTS_PATH.as_posix()}')
            """
        ).fetchone()
        expected_events = conn.execute(
            f"""
            select count(*)
            from read_parquet('{WORKFLOW_PATH.as_posix()}')
            where col2_descr = 'Fakturanr'
              and trim(coalesce(col2_value, '')) <> ''
            """
        ).fetchone()[0]
        monthly_summary = conn.execute(
            f"""
            select
              count(distinct periode) as perioder,
              min(periode) as min_periode,
              max(periode) as max_periode,
              count(distinct substr(periode, 1, 4)) as aar,
              count(distinct omfang_id) filter (where omfang = 'Seksjon') as seksjoner,
              count(*) filter (where omfang = 'Nkom' and kategori = 'Driftskostnader') as nkom_finansieringer,
              count(*) filter (
                where budsjettversjon <> substr(periode, 1, 4) || 'B'
              ) as feil_budsjettversjon
            from read_parquet('{MONTHLY_SUMMARY_PATH.as_posix()}')
            """
        ).fetchone()
        monthly_invoice_errors = conn.execute(
            f"""
            select count(*)
            from read_parquet('{MONTHLY_INVOICES_PATH.as_posix()}')
            where maanedsavslutningsstatus <> 'Kandidat til kontroll'
               or statusgrunnlag is null
               or bokforingskontroll <> 'Ikke bokført i mottatt hovedbokssnapshot'
               or konto is null
               or seksjon is null
               or finansiering is null
            """
        ).fetchone()[0]
        monthly_validation_count = conn.execute(
            f"select count(*) from read_parquet('{MONTHLY_VALIDATION_PATH.as_posix()}')"
        ).fetchone()[0]
        expected_period = conn.execute(
            f"""
            select max(periode)
            from (
              select
                trim(period) as periode,
                max(try_cast(trans_date as date)) as siste_transaksjonsdato,
                count(*) filter (
                  where try_cast(account as integer) between {RULES.account_categories['Lønnskostnader'].first} and {RULES.account_categories['Lønnskostnader'].last}
                ) as lonnsrader
              from read_parquet('{LEDGER_PATH.as_posix()}')
              where regexp_matches(trim(period), '^20[0-9]{{2}}(0[1-9]|1[0-2])$')
              group by trim(period)
            )
            where lonnsrader > 0
              and siste_transaksjonsdato >= last_day(strptime(periode || '01', '%Y%m%d'))
            """
        ).fetchone()[0]
    finally:
        conn.close()

    fakturaer, unique, missing, invalid_quality, safe, ambiguous, unmatched = summary
    if fakturaer == 0:
        raise SystemExit("Workflowrapporten inneholder ingen fakturaer")
    if fakturaer != unique:
        raise SystemExit("Workflowrapporten skal ha nøyaktig én rad per fakturanummer")
    if missing:
        raise SystemExit(f"{missing} rapportlinjer mangler fakturanummer")
    if invalid_quality:
        raise SystemExit(f"{invalid_quality} rader har ukjent koblingskvalitet")
    if invalid_safe:
        raise SystemExit(f"{invalid_safe} rader er feilaktig merket som sikker kobling")
    if safe == 0:
        raise SystemExit("Ingen sikre regnskapskoblinger ble funnet")
    expected_workflow_validations = 8 + sum(
        source is not None
        for source in (
            SOURCES.invoice_queue_history,
            SOURCES.ledger_map,
            SOURCES.receivables,
        )
    )
    if validation_count != expected_workflow_validations:
        raise SystemExit(
            f"Forventet {expected_workflow_validations} workflowkontroller, "
            f"fikk {validation_count}"
        )
    if metadata_rows != 1 or metadata_errors:
        raise SystemExit("Kildestatus skal ha én komplett metadatarad")
    events, event_invoices, events_without_oid, events_without_time, events_without_labels = event_summary
    if events != expected_events:
        raise SystemExit(
            f"Hendelsesvisningen har {events} rader, forventet {expected_events}"
        )
    if event_invoices != fakturaer:
        raise SystemExit(
            f"Hendelsesvisningen dekker {event_invoices} av {fakturaer} fakturaer"
        )
    if events_without_oid:
        raise SystemExit(f"{events_without_oid} hendelsesrader mangler workflow-oid")
    if events_without_time:
        raise SystemExit(f"{events_without_time} hendelsesrader mangler hendelsestid")
    if events_without_labels:
        raise SystemExit(f"{events_without_labels} hendelsesrader mangler forklaring")

    periods, min_period, period, years, sections, nkom_financings, wrong_budget_version = monthly_summary
    if period != expected_period or periods < 1:
        raise SystemExit(
            f"Månedsavslutningen skal bruke siste periode {expected_period}, fikk {period}"
        )
    if SOURCES.cash_ledger and (min_period != "202401" or years != 3):
        raise SystemExit(
            f"Fleirårsvisningen skal dekke 202401–{expected_period}, fikk {min_period}–{period}"
        )
    if sections != len(RULES.sections):
        raise SystemExit(f"Forventet {len(RULES.sections)} seksjoner fra malen, fikk {sections}")
    if nkom_financings == 0:
        raise SystemExit("Mangler Nkom-totaler per finansiering")
    if wrong_budget_version:
        raise SystemExit(
            "Månedsavslutningen bruker feil budsjettversjon for eitt eller fleire år"
        )
    if monthly_invoice_errors:
        raise SystemExit(f"{monthly_invoice_errors} fakturarader bryter månedsavslutningsreglene")
    if monthly_validation_count != 8:
        raise SystemExit(
            f"Forventet 8 månedsavslutningskontroller, fikk {monthly_validation_count}"
        )

    period_workbook = ROOT / "static" / f"manedsavslutning_{period[:4]}-{period[4:]}.xlsx"
    if not period_workbook.exists():
        raise SystemExit(f"Mangler periodeversjon av Excel-filen: {period_workbook.name}")
    workbook = load_workbook(period_workbook, data_only=False, read_only=True)
    if workbook[f"Totalt eks {RULES.cash.section}"]["C7"].value in (None, ""):
        raise SystemExit("Excel-filen har blank hovedverdi i totalfanen")
    if workbook["711 - SID"]["C26"].data_type == "f":
        raise SystemExit("Excel-filen bruker fortsatt en formel uten forhåndsverdi i C26")
    expected_cash = duckdb.sql(
        f"""
        select sum(try_cast({'cash_amount' if SOURCES.cash_ledger else 'amount'} as double))
        from read_parquet('{(SOURCES.cash_ledger or LEDGER_PATH).as_posix()}')
        where trim(dim_1) = '{RULES.cash.section}'
          and trim(account) = '{RULES.cash.account}'
          and trim(dim_4) = '{RULES.cash.financing}'
          and trim({'pay_period' if SOURCES.cash_ledger else 'period'})
              between '{period[:4]}01' and '{period}'
        """
    ).fetchone()[0]
    if workbook[RULES.cash.section]["F6"].value != expected_cash:
        raise SystemExit(
            f"Excel-filen mangler hittil-i-år for konto {RULES.cash.account} "
            f"i fane {RULES.cash.section}"
        )
    cash_detail = workbook[f"{RULES.cash.section} kontantdetaljer"]
    if cash_detail["G" + str(cash_detail.max_row)].value != expected_cash:
        raise SystemExit("Detaljfanen for 712 avstemmer ikke mot kontantkilden")

    print("Oppgave 3-validering bestått")
    print(f"- {fakturaer} unike workflowfakturaer")
    print(f"- {events} oppgavehendelser med workflow-oid")
    print(f"- {safe} sikre koblinger")
    print(f"- {ambiguous} tvetydige koblinger")
    print(f"- {unmatched} uten regnskapskobling")
    print(f"- Månedsavslutning for periode {period} med {sections} seksjoner")
    print(f"- Utfylt Excel-mal: {MONTHLY_WORKBOOK_PATH.name}")


if __name__ == "__main__":
    main()
