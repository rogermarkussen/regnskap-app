from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import load_workbook

try:
    from .project_data import task2_sources
    from .parquet_report import INVESTMENT_ACCOUNTS, ParquetReportSources, build_parquet_report
except ImportError:
    from project_data import task2_sources
    from parquet_report import INVESTMENT_ACCOUNTS, ParquetReportSources, build_parquet_report


CODE_ROOT = Path(__file__).resolve().parents[1]
ROOT = CODE_ROOT.parent
SOURCES = task2_sources()
SOURCE_DATA_DIR = SOURCES.dashboard_workbook.parent
PARQUET_DIR = SOURCES.generated_dir / "static-app"

SOURCE_DASHBOARD_WORKBOOK = SOURCES.dashboard_workbook
GROUPING_WORKBOOK = SOURCES.account_grouping_workbook
RAW_TRANSACTION_WORKBOOK = SOURCES.raw_transactions_workbook
BUDGET_HEADER_PARQUET = SOURCES.budget_header
BUDGET_VALUE_PARQUET = SOURCES.budget_values
DIMENSION_VALUES_PARQUET = SOURCES.dimension_values
LEDGER_PARQUET = SOURCES.ledger
BUDGET_VERSION = "2026B"


@dataclass(frozen=True)
class SectionScope:
    code: str
    name: str
    label: str
    sort_order: int

FINANCE_SHEETS = {
    "Finansiering 154301": "154301",
    "Finansiering 154345": "154345",
    "Fin 154322 og 045101": "154322+045101",
}


def slug(value: object) -> str:
    text = "" if value is None else str(value)
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", normalized).strip("_").lower()
    return normalized or "unnamed"


def clean_cell(value: object) -> object:
    if isinstance(value, str):
        value = value.replace("\xa0", " ").strip()
        return None if value == "" else value
    return value


def account_parts(value: object) -> tuple[str | None, str | None]:
    if not isinstance(value, str):
        return None, None
    match = re.match(r"^(\d{4})\s*-\s*(.+)$", value.strip())
    if not match:
        return None, None
    return match.group(1), match.group(2).strip()


def write_parquet(conn: duckdb.DuckDBPyConnection, name: str, df: pd.DataFrame) -> Path:
    path = PARQUET_DIR / f"{name}.parquet"
    conn.register("df_to_write", df)
    conn.execute(
        f"""
        copy df_to_write
        to '{path.as_posix()}'
        (format parquet, compression zstd)
        """
    )
    conn.unregister("df_to_write")
    return path


def workbook_raw_tables(conn: duckdb.DuckDBPyConnection) -> list[tuple[str, Path]]:
    outputs: list[tuple[str, Path]] = []
    # Manifestet peker denne mappen til de midlertidige operative Excel-kildene.
    for workbook_path in sorted(SOURCE_DATA_DIR.glob("*.xlsx")):
        if workbook_path.name.startswith("~$"):
            continue
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [clean_cell(value) for value in row]
                if not any(value is not None for value in values):
                    continue
                rows.append(
                    {
                        "source_file": workbook_path.name,
                        "sheet_name": sheet.title,
                        "excel_row": row_number,
                        **{f"col_{index + 1:02d}": value for index, value in enumerate(values)},
                    }
                )
            if rows:
                name = f"raw_{slug(workbook_path.stem)}_{slug(sheet.title)}"
                outputs.append((name, write_parquet(conn, name, pd.DataFrame(rows))))
    return outputs


def finance_rows_frame(workbook_path: Path = SOURCE_DASHBOARD_WORKBOOK) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []
    for sheet_name, financing in FINANCE_SHEETS.items():
        worksheet = workbook[sheet_name]
        for excel_row, row in enumerate(worksheet.iter_rows(min_row=9, values_only=True), start=9):
            values = [clean_cell(value) for value in row]
            label = values[0] if values else None
            if not label:
                continue
            account_number, account_name = account_parts(label)
            numeric_values = [value for value in values[2:25] if isinstance(value, int | float)]
            row_type = "account" if account_number else ("subtotal" if numeric_values else "section")
            rows.append(
                {
                    "finansiering": financing,
                    "sheet_name": sheet_name,
                    "excel_row": excel_row,
                    "row_type": row_type,
                    "konto_text": label,
                    "konto": account_number,
                    "konto_navn": account_name,
                    "virksomhet_budsjett_tusen": values[2] if len(values) > 2 else None,
                    "hovedbok_tusen": values[3] if len(values) > 3 else None,
                    "avvik_tusen": values[4] if len(values) > 4 else None,
                    "aarets_budsjett_tusen": values[7] if len(values) > 7 else None,
                    "forbruk_av_aarets_budsjett_pct": values[8] if len(values) > 8 else None,
                    **{
                        f"periode_{period}_tusen": values[index]
                        for index, period in enumerate(range(202601, 202613), start=9)
                        if len(values) > index
                    },
                    "kontant_budsjett_tusen": values[22] if len(values) > 22 else None,
                    "kontant_tusen": values[23] if len(values) > 23 else None,
                    "kontant_avvik_tusen": values[24] if len(values) > 24 else None,
                }
            )
    return pd.DataFrame(rows)


def account_groups(conn: duckdb.DuckDBPyConnection) -> Path:
    rows = [
        {
            "hovedgruppe": group["hovedgruppe"],
            "undergruppe": group["undergruppe"],
            "konto_text": account["konto_text"],
            "konto": account["konto"],
            "konto_navn": account["konto_navn"],
            "excel_row": excel_row,
        }
        for excel_row, group in enumerate(grouping_structure(), start=1)
        for account in group["accounts"]
    ]
    return write_parquet(conn, "account_groups", pd.DataFrame(rows))


REPORT_VALUE_COLUMNS = [
    "virksomhet_budsjett_tusen",
    "hovedbok_tusen",
    "avvik_tusen",
    "aarets_budsjett_tusen",
    *[f"budsjett_{period}_tusen" for period in range(202601, 202613)],
    "kontant_budsjett_tusen",
    "kontant_tusen",
    "kontant_avvik_tusen",
]
SUMMED_VALUE_COLUMNS = REPORT_VALUE_COLUMNS

def latest_complete_ledger_period(
    ledger_path: Path = LEDGER_PARQUET,
) -> int:
    """Finn siste måned der hovedboken inneholder bilag datert til månedsslutt."""
    if not ledger_path.exists():
        raise FileNotFoundError(f"Mangler operativ hovedbok: {ledger_path.name}")

    conn = duckdb.connect()
    try:
        row = conn.execute(
            f"""
            select try_cast(period as integer) as period
            from read_parquet('{ledger_path.as_posix()}')
            where try_cast(period as integer) between 202601 and 202612
            group by period
            having max(try_cast(voucher_date as date)) >= last_day(
              strptime(cast(try_cast(period as integer) as varchar) || '01', '%Y%m%d')
            )
            order by period desc
            limit 1
            """
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        raise ValueError("Fant ingen komplett hovedboksperiode i 2026")
    return int(row[0])


def period_label(period_to: int) -> str:
    return f"01–{period_to % 100:02d} {period_to // 100}"


LATEST_COMPLETE_PERIOD = latest_complete_ledger_period()

TASK2_PERIODS = {
    "p1_3": (202603, "01–03 2026"),
    "p1_4": (202604, "01–04 2026"),
    "p1_6": (202606, "01–06 2026"),
    "latest": (LATEST_COMPLETE_PERIOD, period_label(LATEST_COMPLETE_PERIOD)),
}


def grouping_structure() -> list[dict[str, object]]:
    worksheet = load_workbook(GROUPING_WORKBOOK, read_only=True, data_only=True)["Kontogruppering"]
    groups: list[dict[str, object]] = []
    current_main_group: str | None = None
    seen_accounts: set[str] = set()

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        values = [clean_cell(value) for value in row]
        if len(values) > 1 and values[1]:
            current_main_group = str(values[1])
        subgroup = values[2] if len(values) > 2 else None
        if not current_main_group or not subgroup:
            continue

        accounts_by_number: dict[str, dict[str, str]] = {}
        for cell in values[3:]:
            account_number, account_name = account_parts(cell)
            if not account_number:
                continue
            # 5405 står dobbelt i kildefilen. Behold én rad og den siste
            # teksten, slik at samme konto aldri summeres to ganger.
            accounts_by_number[account_number] = {
                "konto": account_number,
                "konto_navn": account_name or "",
                "konto_text": str(cell),
            }

        accounts = []
        for account_number, account in accounts_by_number.items():
            if account_number in seen_accounts:
                continue
            seen_accounts.add(account_number)
            accounts.append(account)
        groups.append(
            {
                "hovedgruppe": current_main_group,
                "undergruppe": str(subgroup),
                "accounts": accounts,
            }
        )

    conn = duckdb.connect()
    try:
        investment_names = dict(
            conn.execute(
                f"""
                select lpad(trim(dim_value), 4, '0') as konto,
                  any_value(trim(description)) as konto_navn
                from read_parquet('{DIMENSION_VALUES_PARQUET.as_posix()}')
                where attribute_id = 'A0'
                  and lpad(trim(dim_value), 4, '0') in ('1250', '1270', '1280', '1281')
                group by 1
                """
            ).fetchall()
        )
    finally:
        conn.close()
    investment_accounts = [
        {
            "konto": account,
            "konto_navn": investment_names.get(account, "Kontonavn mangler"),
            "konto_text": f"{account} - {investment_names.get(account, 'Kontonavn mangler')}",
        }
        for account in INVESTMENT_ACCOUNTS
    ]
    groups.insert(
        0,
        {
            "hovedgruppe": "Investeringsrapport",
            "undergruppe": "Varige driftsmidler",
            "accounts": investment_accounts,
        },
    )
    return groups


def raw_actuals_frame() -> pd.DataFrame:
    worksheet = load_workbook(RAW_TRANSACTION_WORKBOOK, read_only=True, data_only=True)["Ark1"]
    totals: dict[tuple[str, str], float] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        period = row[5] if len(row) > 5 else None
        account = row[7] if len(row) > 7 else None
        financing = row[15] if len(row) > 15 else None
        amount = row[26] if len(row) > 26 else None
        if not isinstance(period, int | float) or not 202601 <= int(period) <= 202603:
            continue
        if not isinstance(account, int | float) or not isinstance(amount, int | float):
            continue
        account_text = str(int(account)).zfill(4)
        financing_text = "" if financing is None else str(int(financing)) if isinstance(financing, int | float) else str(financing)
        totals[(financing_text, account_text)] = totals.get((financing_text, account_text), 0.0) + float(amount) / 1000

    return pd.DataFrame(
        [
            {"finansiering": financing, "konto": account, "hovedbok_tusen": amount}
            for (financing, account), amount in totals.items()
        ]
    )


def _sum_by_account(df: pd.DataFrame, columns: list[str]) -> dict[str, dict[str, float | None]]:
    if df.empty:
        return {}
    grouped = df.groupby("konto", as_index=False)[columns].sum(min_count=1)
    return {
        str(row["konto"]): {
            column: None if pd.isna(row[column]) else float(row[column])
            for column in columns
        }
        for _, row in grouped.iterrows()
    }


def _cash_values_for_account(
    cash: dict[str, dict[str, float | None]],
    account: str,
    *,
    available: bool,
) -> tuple[float | None, float | None]:
    """Behold eksplisitt null, men ikke erstatt manglende kontantdata med null."""
    if not available or account not in cash:
        return None, None
    account_cash = cash[account]
    return (
        account_cash.get("kontant_budsjett_tusen"),
        account_cash.get("kontant_tusen"),
    )


def synapse_budget_by_account(
    *,
    dim_1: str | None = None,
    dim_2: str | None = None,
    exclude_dim_1: tuple[str, ...] = (),
    section_code: str | None = None,
) -> dict[str, dict[str, float]]:
    """Returner operativt 2026B-budsjett per konto og måned for ett rapportvalg."""
    missing = [
        path.name
        for path in (BUDGET_HEADER_PARQUET, BUDGET_VALUE_PARQUET)
        if not path.exists()
    ]
    if missing:
        raise FileNotFoundError(f"Mangler operative budsjettfiler: {', '.join(missing)}")

    conn = duckdb.connect()
    try:
        filters = ["h.version = ?", "try_cast(v.period as integer) between 202601 and 202612"]
        parameters: list[object] = [BUDGET_VERSION]
        if dim_1 is not None:
            filters.append("trim(h.dim_1) = ?")
            parameters.append(dim_1)
        if dim_2 is not None:
            filters.append("trim(h.dim_2) = ?")
            parameters.append(dim_2)
        if exclude_dim_1:
            placeholders = ", ".join("?" for _ in exclude_dim_1)
            filters.append(f"(h.dim_1 is null or trim(h.dim_1) not in ({placeholders}))")
            parameters.extend(exclude_dim_1)
        if section_code is not None:
            filters.append("trim(h.dim_1) = ?")
            parameters.append(section_code)
        budget_rows = conn.execute(
            f"""
            select
              lpad(cast(h.account as varchar), 4, '0') as konto,
              cast(v.period as integer) as period,
              sum(try_cast(v.amount as double)) / 1000 as amount_tusen
            from read_parquet('{BUDGET_HEADER_PARQUET.as_posix()}') h
            join read_parquet('{BUDGET_VALUE_PARQUET.as_posix()}') v using (trans_id)
            where {' and '.join(filters)}
            group by h.account, v.period
            """,
            parameters,
        ).df()
    finally:
        conn.close()

    result: dict[str, dict[str, float]] = {}
    for row in budget_rows.itertuples(index=False):
        result.setdefault(str(row.konto), {})[
            f"periode_{int(row.period)}_tusen"
        ] = float(row.amount_tusen)
    return result


def parquet_actuals_by_account(
    *,
    financings: tuple[str, ...],
    period_to: int,
    project: str | None = None,
    section_code: str | None = None,
) -> dict[str, dict[str, float]]:
    """Summer hovedbok fra kanonisk Parquet uten å lese Excel."""
    if not LEDGER_PARQUET.exists():
        raise FileNotFoundError(f"Mangler operativ hovedbok: {LEDGER_PARQUET.name}")
    placeholders = ", ".join("?" for _ in financings)
    filters = [
        f"trim(dim_4) in ({placeholders})",
        "try_cast(period as integer) between 202601 and ?",
    ]
    parameters: list[object] = [*financings, period_to]
    if project is not None:
        filters.append("trim(dim_2) = ?")
        parameters.append(project)
    if section_code is not None:
        filters.append("trim(dim_1) = ?")
        parameters.append(section_code)
    conn = duckdb.connect()
    try:
        frame = conn.execute(
            f"""
            select
              lpad(trim(account), 4, '0') as konto,
              sum(try_cast(amount as double)) / 1000 as hovedbok_tusen
            from read_parquet('{LEDGER_PARQUET.as_posix()}')
            where {' and '.join(filters)}
            group by account
            """,
            parameters,
        ).df()
    finally:
        conn.close()
    return _sum_by_account(frame, ["hovedbok_tusen"])


def calculated_account_values(section_code: str | None = None) -> dict[
    tuple[str, str], dict[str, dict[str, float | None]]
]:
    finance = None
    snapshot_actuals = None
    if section_code is None:
        finance = finance_rows_frame(SOURCE_DASHBOARD_WORKBOOK)
        finance = finance[finance["row_type"] == "account"].copy()
        snapshot_actuals = raw_actuals_frame()
    all_financing_budget = synapse_budget_by_account(section_code=section_code)
    mapped_154301_budget = synapse_budget_by_account(
        exclude_dim_1=("212", "761"), section_code=section_code
    )
    source_months = [f"periode_{period}_tusen" for period in range(202601, 202613)]
    result: dict[tuple[str, str], dict[str, dict[str, float | None]]] = {}

    rules = {
        "154301": {
            "budget_financing": ["154301"],
            "cash_financing": ["154301"],
            "cash_period": "p1_3",
            "parquet_budget": mapped_154301_budget,
            "actual_financings": ("154301",),
        },
        "154345": {
            "budget_financing": ["154345"],
            "cash_financing": ["154345"],
            "cash_period": "p1_4",
            "parquet_budget": synapse_budget_by_account(
                dim_1="212", section_code=section_code
            ),
            "actual_financings": ("154345",),
        },
        "154322+045101": {
            "budget_financing": ["154322+045101"],
            "cash_financing": ["154322+045101"],
            "cash_period": "p1_3",
            "parquet_budget": synapse_budget_by_account(
                dim_1="761", section_code=section_code
            ),
            "actual_financings": ("154322", "045101"),
        },
        "alle": {
            "budget_financing": ["154301", "154345", "154322+045101"],
            # 154345-arket gjelder januar-april. Kontantkolonnen kan derfor
            # ikke tas med i en januar-mars-rapport.
            "cash_financing": ["154301", "154322+045101"],
            "cash_period": "p1_3",
            "parquet_budget": all_financing_budget,
            "actual_financings": ("154301", "154322", "045101", "154345"),
        },
    }

    for report_financing, rule in rules.items():
        snapshot_budget: dict[str, dict[str, float | None]] = {}
        snapshot_cash: dict[str, dict[str, float | None]] = {}
        if finance is not None:
            budget_rows = finance[finance["finansiering"].isin(rule["budget_financing"])]
            cash_rows = finance[finance["finansiering"].isin(rule["cash_financing"])]
            snapshot_budget = _sum_by_account(budget_rows, source_months)
            snapshot_cash = _sum_by_account(
                cash_rows, ["kontant_budsjett_tusen", "kontant_tusen"]
            )

        for period_key, (period_to, _) in TASK2_PERIODS.items():
            if (
                section_code is None
                and period_key == "p1_3"
                and report_financing == "154301"
            ):
                budget = snapshot_budget
                assert snapshot_actuals is not None
                actual_rows = snapshot_actuals[
                    snapshot_actuals["finansiering"].isin(["154301"])
                ]
                actual = _sum_by_account(actual_rows, ["hovedbok_tusen"])
            elif (
                section_code is None
                and period_key == "p1_3"
                and report_financing == "alle"
            ):
                budget = all_financing_budget
                assert snapshot_actuals is not None
                actual = _sum_by_account(snapshot_actuals, ["hovedbok_tusen"])
            else:
                budget = rule["parquet_budget"]
                actual = parquet_actuals_by_account(
                    financings=rule["actual_financings"],
                    period_to=period_to,
                    project=rule.get("project"),
                    section_code=section_code,
                )

            cash_available = section_code is None and period_key == rule["cash_period"]
            cash = snapshot_cash if cash_available else {}
            accounts = set(budget) | set(cash) | set(actual)
            values_by_account: dict[str, dict[str, float | None]] = {}
            for account in accounts:
                monthly = {
                    f"budsjett_{period}_tusen": budget.get(account, {}).get(
                        f"periode_{period}_tusen"
                    )
                    for period in range(202601, 202613)
                }
                period_values = [
                    monthly[f"budsjett_{period}_tusen"]
                    for period in range(202601, period_to + 1)
                    if monthly[f"budsjett_{period}_tusen"] is not None
                ]
                annual_values = [value for value in monthly.values() if value is not None]
                period_budget = sum(period_values) if period_values else None
                annual_budget = sum(annual_values) if annual_values else None
                actual_value = actual.get(account, {}).get("hovedbok_tusen") or 0.0
                cash_budget, cash_value = _cash_values_for_account(
                    cash,
                    account,
                    available=cash_available,
                )
                values_by_account[account] = {
                    "virksomhet_budsjett_tusen": period_budget,
                    "hovedbok_tusen": actual_value,
                    "avvik_tusen": (
                        None if period_budget is None else period_budget - actual_value
                    ),
                    "aarets_budsjett_tusen": annual_budget,
                    **monthly,
                    "kontant_budsjett_tusen": cash_budget,
                    "kontant_tusen": cash_value,
                    "kontant_avvik_tusen": (
                        None
                        if cash_budget is None or cash_value is None
                        else cash_budget - cash_value
                    ),
                }
            result[(report_financing, period_key)] = values_by_account
    return result


def _summed_values(rows: list[dict[str, object]]) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for column in SUMMED_VALUE_COLUMNS:
        present = [row.get(column) for row in rows if row.get(column) is not None]
        values[column] = sum(float(value) for value in present) if present else None
    missing_period_budget = any(
        row.get("virksomhet_budsjett_tusen") is None
        and row.get("hovedbok_tusen") not in (None, 0, 0.0)
        for row in rows
    )
    missing_annual_budget = any(
        row.get("aarets_budsjett_tusen") is None
        and row.get("hovedbok_tusen") not in (None, 0, 0.0)
        for row in rows
    )
    if missing_period_budget:
        values["avvik_tusen"] = None
    annual = values["aarets_budsjett_tusen"]
    actual = values["hovedbok_tusen"]
    values["forbruk_av_aarets_budsjett"] = (
        None
        if missing_annual_budget or annual in (None, 0) or actual is None
        else float(actual) / float(annual)
    )
    return values


def _empty_account_values() -> dict[str, float | None]:
    values: dict[str, float | None] = {
        column: 0.0 for column in SUMMED_VALUE_COLUMNS
    }
    values.update(
        {
            "kontant_budsjett_tusen": None,
            "kontant_tusen": None,
            "kontant_avvik_tusen": None,
        }
    )
    return values


def section_scopes() -> tuple[SectionScope, ...]:
    missing = [
        path.name
        for path in (LEDGER_PARQUET, BUDGET_HEADER_PARQUET, DIMENSION_VALUES_PARQUET)
        if not path.exists()
    ]
    if missing:
        raise FileNotFoundError(f"Mangler kilder for seksjonsfilter: {', '.join(missing)}")

    conn = duckdb.connect()
    try:
        rows = conn.execute(
            f"""
            with relevant_codes as (
              select distinct trim(dim_1) as code
              from read_parquet('{LEDGER_PARQUET.as_posix()}')
              where trim(dim_4) in ('154301', '154345', '154322', '045101')
                and trim(dim_1) <> ''
              union
              select distinct trim(dim_1) as code
              from read_parquet('{BUDGET_HEADER_PARQUET.as_posix()}')
              where version = '{BUDGET_VERSION}'
                and trim(dim_1) <> ''
            ), names as (
              select
                trim(dim_value) as code,
                any_value(trim(description)) as name
              from read_parquet('{DIMENSION_VALUES_PARQUET.as_posix()}')
              where attribute_id = 'C1'
              group by 1
            )
            select
              relevant_codes.code,
              coalesce(names.name, 'Navn mangler i dimensjonsregisteret') as name
            from relevant_codes
            left join names using (code)
            order by try_cast(relevant_codes.code as integer), relevant_codes.code
            """
        ).fetchall()
    finally:
        conn.close()

    scopes = [SectionScope("all", "Alle seksjoner", "Alle seksjoner", 0)]
    for code, name in rows:
        sort_order = int(code) if str(code).isdigit() else 90_000
        scopes.append(SectionScope(str(code), str(name), f"{code} · {name}", sort_order))
    return tuple(scopes)


def grouped_finance_rows_frame(
    scope: SectionScope | None = None,
) -> pd.DataFrame:
    selected_scope = scope or SectionScope("all", "Alle seksjoner", "Alle seksjoner", 0)
    structure = grouping_structure()
    account_values = calculated_account_values(
        None if selected_scope.code == "all" else selected_scope.code
    )
    rows: list[dict[str, object]] = []

    report_options = [
        ("154301", "Finansiering 154301"),
        ("154345", "Finansiering 154345"),
        ("154322+045101", "Finansiering 154322 + 045101"),
        ("alle", "Alle finansieringer"),
    ]
    cash_periods = {
        "154301": "p1_3",
        "154345": "p1_4",
        "154322+045101": "p1_3",
        "alle": "p1_3",
    }
    for financing, financing_label, period_key in [
        (financing, label, period_key)
        for financing, label in report_options
        for period_key in TASK2_PERIODS
    ]:
        _, periodetekst = TASK2_PERIODS[period_key]
        source_files = (
            f"{LEDGER_PARQUET.name}; {BUDGET_HEADER_PARQUET.name}; "
            f"{BUDGET_VALUE_PARQUET.name} ({BUDGET_VERSION})"
        )
        if (
            selected_scope.code == "all"
            and period_key == "p1_3"
            and financing in {"154301", "alle"}
        ):
            source_files = f"{SOURCE_DASHBOARD_WORKBOOK.name}; {RAW_TRANSACTION_WORKBOOK.name}"
            if financing == "alle":
                source_files += f"; {BUDGET_HEADER_PARQUET.name}; {BUDGET_VALUE_PARQUET.name} ({BUDGET_VERSION})"
        if selected_scope.code == "all" and period_key == (
            "p1_4" if financing == "154345" else "p1_3"
        ):
            source_files += f"; {SOURCE_DASHBOARD_WORKBOOK.name} (kun kontant)"
        scope_fields = {
            "section_code": selected_scope.code,
            "section_name": selected_scope.name,
            "section_label": selected_scope.label,
            "section_sort": selected_scope.sort_order,
        }
        excel_row = 1
        report_rows: list[dict[str, object]] = []
        main_groups = list(dict.fromkeys(str(group["hovedgruppe"]) for group in structure))
        for main_group in main_groups:
            report_rows.append(
                {
                    **scope_fields,
                    "finansiering": financing,
                    "finansiering_tekst": financing_label,
                    "rapportperiode": period_key,
                    "periodetekst": periodetekst,
                    "excel_row": excel_row,
                    "hovedgruppe": main_group,
                    "row_type": "section",
                    "radtekst": main_group,
                    "konto": None,
                    "konto_navn": None,
                    **{column: None for column in SUMMED_VALUE_COLUMNS},
                    "forbruk_av_aarets_budsjett": None,
                    "source_file": source_files,
                }
            )
            excel_row += 1
            main_group_accounts: list[dict[str, object]] = []
            for group in [candidate for candidate in structure if candidate["hovedgruppe"] == main_group]:
                group_accounts: list[dict[str, object]] = []
                for account in group["accounts"]:
                    account_number = str(account["konto"])
                    has_operational_data = account_number in account_values[(financing, period_key)]
                    values = (
                        account_values[(financing, period_key)][account_number]
                        if has_operational_data
                        else _empty_account_values()
                    )
                    if period_key != cash_periods[financing]:
                        values = {
                            **values,
                            "kontant_budsjett_tusen": None,
                            "kontant_tusen": None,
                            "kontant_avvik_tusen": None,
                        }
                    account_row = {
                        **scope_fields,
                        "finansiering": financing,
                        "finansiering_tekst": financing_label,
                        "rapportperiode": period_key,
                        "periodetekst": periodetekst,
                        "excel_row": 0,
                        "hovedgruppe": main_group,
                        "row_type": "account",
                        "radtekst": account["konto_text"],
                        "konto": account_number,
                        "konto_navn": account["konto_navn"],
                        "data_status": (
                            "Operative tall" if has_operational_data else "Ingen operative tall"
                        ),
                        **values,
                        "forbruk_av_aarets_budsjett": (
                            None
                            if values["aarets_budsjett_tusen"] in (None, 0) or values["hovedbok_tusen"] is None
                            else float(values["hovedbok_tusen"]) / float(values["aarets_budsjett_tusen"])
                        ),
                        "source_file": source_files,
                    }
                    group_accounts.append(account_row)
                    main_group_accounts.append(account_row)
                if not group_accounts:
                    continue
                report_rows.append(
                    {
                        **scope_fields,
                        "finansiering": financing,
                        "finansiering_tekst": financing_label,
                        "rapportperiode": period_key,
                        "periodetekst": periodetekst,
                        "excel_row": excel_row,
                        "hovedgruppe": main_group,
                        "row_type": "group",
                        "radtekst": group["undergruppe"],
                        "konto": None,
                        "konto_navn": None,
                        **_summed_values(group_accounts),
                        "source_file": source_files,
                    }
                )
                excel_row += 1
                for account_row in group_accounts:
                    account_row["excel_row"] = excel_row
                    report_rows.append(account_row)
                    excel_row += 1

            report_rows.append(
                {
                    **scope_fields,
                    "finansiering": financing,
                    "finansiering_tekst": financing_label,
                    "rapportperiode": period_key,
                    "periodetekst": periodetekst,
                    "excel_row": excel_row,
                    "hovedgruppe": main_group,
                    "row_type": "total",
                    "radtekst": (
                        "Totale investeringer"
                        if main_group == "Investeringsrapport"
                        else f"Totale {main_group.lower()}"
                    ),
                    "konto": None,
                    "konto_navn": None,
                    **_summed_values(main_group_accounts),
                    "source_file": source_files,
                }
            )
            excel_row += 1

        all_accounts = [
            row
            for row in report_rows
            if row["row_type"] == "account"
            and row["hovedgruppe"] != "Investeringsrapport"
        ]
        report_rows.append(
            {
                **scope_fields,
                "finansiering": financing,
                "finansiering_tekst": financing_label,
                "rapportperiode": period_key,
                "periodetekst": periodetekst,
                "excel_row": excel_row,
                "hovedgruppe": None,
                "row_type": "total",
                "radtekst": "Driftskostnader",
                "konto": None,
                "konto_navn": None,
                **_summed_values(all_accounts),
                "source_file": source_files,
            }
        )
        rows.extend(report_rows)
    return pd.DataFrame(rows)


def grouped_finance_rows(conn: duckdb.DuckDBPyConnection) -> Path:
    return write_parquet(conn, "grouped_finance_rows", grouped_finance_rows_frame())


def section_grouped_finance_rows(conn: duckdb.DuckDBPyConnection) -> Path:
    frames = [grouped_finance_rows_frame(scope) for scope in section_scopes()[1:]]
    return write_parquet(conn, "section_grouped_finance_rows", pd.concat(frames, ignore_index=True))


def write_static_app_data(
    grouped_path: Path,
    section_grouped_path: Path,
) -> Path:
    PARQUET_DIR.mkdir(parents=True, exist_ok=True)
    output_path = PARQUET_DIR / "task2-report.parquet"
    conn = duckdb.connect()
    try:
        conn.execute(
            f"""
            copy (
              select * from read_parquet('{grouped_path.as_posix()}')
              union all by name
              select * from read_parquet('{section_grouped_path.as_posix()}')
              order by section_sort, finansiering, rapportperiode, excel_row
            ) to '{output_path.as_posix()}'
            (format parquet, compression zstd)
            """
        )
    finally:
        conn.close()
    return output_path


def grouped_finance_validation(conn: duckdb.DuckDBPyConnection) -> Path:
    grouped_path = PARQUET_DIR / "grouped_finance_rows.parquet"
    grouping_path = PARQUET_DIR / "account_groups.parquet"
    grouped = conn.execute(f"select * from read_parquet('{grouped_path.as_posix()}')").df()
    expected = conn.execute(
        f"select distinct konto from read_parquet('{grouping_path.as_posix()}') where konto is not null"
    ).df()["konto"].astype(str)
    expected_accounts = set(expected)
    results = []

    def add(
        financing: str,
        report_period: str,
        check: str,
        status: str,
        detail: str,
        discrepancies: int,
    ) -> None:
        results.append(
            {
                "finansiering": financing,
                "rapportperiode": report_period,
                "kontroll": check,
                "status": status,
                "detalj": detail,
                "antall_avvik": discrepancies,
            }
        )

    for (financing, report_period), subset in grouped.groupby(
        ["finansiering", "rapportperiode"], sort=True
    ):
        actual_accounts = set(subset.loc[subset["row_type"] == "account", "konto"].dropna().astype(str))
        missing = sorted(expected_accounts - actual_accounts)
        add(
            financing,
            report_period,
            "Kontodekning",
            "warning" if missing else "ok",
            f"{len(actual_accounts)} av {len(expected_accounts)} kontoer. Mangler i tallrapporten: {', '.join(missing)}"
            if missing
            else f"Alle {len(expected_accounts)} kontoer er med.",
            len(missing),
        )

        arithmetic = subset.dropna(
            subset=["virksomhet_budsjett_tusen", "hovedbok_tusen", "avvik_tusen"]
        )
        arithmetic_errors = (
            (
                arithmetic["virksomhet_budsjett_tusen"]
                - arithmetic["hovedbok_tusen"]
                - arithmetic["avvik_tusen"]
            ).abs()
            > 1e-6
        ).sum()
        add(
            financing,
            report_period,
            "Budsjett − hovedbok = avvik",
            "ok" if arithmetic_errors == 0 else "error",
            f"{len(arithmetic)} rader kontrollert.",
            int(arithmetic_errors),
        )

        month_columns = [f"budsjett_{period}_tusen" for period in range(202601, 202613)]
        annual = subset.dropna(subset=["aarets_budsjett_tusen", *month_columns])
        annual_errors = (
            (annual[month_columns].sum(axis=1) - annual["aarets_budsjett_tusen"]).abs() > 1e-5
        ).sum()
        add(
            financing,
            report_period,
            "Månedsbudsjett = årsbudsjett",
            "ok" if annual_errors == 0 else "error",
            f"{len(annual)} rader kontrollert.",
            int(annual_errors),
        )

        cash = subset.dropna(
            subset=["kontant_budsjett_tusen", "kontant_tusen", "kontant_avvik_tusen"]
        )
        cash_errors = (
            (
                cash["kontant_budsjett_tusen"]
                - cash["kontant_tusen"]
                - cash["kontant_avvik_tusen"]
            ).abs()
            > 1e-6
        ).sum()
        add(
            financing,
            report_period,
            "Kontantbudsjett − kontant = avvik",
            "ok" if cash_errors == 0 else "error",
            f"{len(cash)} rader kontrollert.",
            int(cash_errors),
        )

        missing_cash_inputs_with_deviation = subset[
            (
                subset["kontant_budsjett_tusen"].isna()
                | subset["kontant_tusen"].isna()
            )
            & subset["kontant_avvik_tusen"].notna()
        ]
        add(
            financing,
            report_period,
            "Manglende kontantgrunnlag gir ikke avvik",
            "ok" if missing_cash_inputs_with_deviation.empty else "error",
            (
                "Manglende kontantbudsjett eller kontantverdi beholdes som tom verdi."
                if missing_cash_inputs_with_deviation.empty
                else (
                    f"{len(missing_cash_inputs_with_deviation)} rader har kontantavvik "
                    "uten komplett kontantgrunnlag."
                )
            ),
            len(missing_cash_inputs_with_deviation),
        )

    return write_parquet(conn, "grouped_finance_validation", pd.DataFrame(results))


def main() -> None:
    PARQUET_DIR.mkdir(parents=True, exist_ok=True)
    for path in PARQUET_DIR.glob("*.parquet"):
        path.unlink()

    temp_db = PARQUET_DIR / "_writer.duckdb"
    if temp_db.exists():
        temp_db.unlink()
    conn = duckdb.connect(temp_db)
    try:
        if SOURCES.cash_ledger and SOURCES.cash_accounts and SOURCES.account_plan:
            groups, report, validations = build_parquet_report(
                ParquetReportSources(
                    ledger=SOURCES.ledger,
                    budget_header=SOURCES.budget_header,
                    budget_values=SOURCES.budget_values,
                    dimension_values=SOURCES.dimension_values,
                    cash_ledger=SOURCES.cash_ledger,
                    cash_accounts=SOURCES.cash_accounts,
                    account_plan=SOURCES.account_plan,
                )
            )
            groups_path = write_parquet(conn, "account_groups", groups)
            report_path = write_parquet(conn, "grouped_finance_rows", report[report.section_code == "all"])
            section_path = write_parquet(conn, "section_grouped_finance_rows", report[report.section_code != "all"])
            validation_path = write_parquet(conn, "grouped_finance_validation", validations)
            parquet_outputs = [
                ("account_groups", groups_path),
                ("grouped_finance_rows", report_path),
                ("section_grouped_finance_rows", section_path),
                ("grouped_finance_validation", validation_path),
            ]
        else:
            parquet_outputs = [
                ("account_groups", account_groups(conn)),
                ("grouped_finance_rows", grouped_finance_rows(conn)),
                ("section_grouped_finance_rows", section_grouped_finance_rows(conn)),
                ("grouped_finance_validation", grouped_finance_validation(conn)),
            ]
    finally:
        conn.close()
        for suffix in ("", ".wal"):
            path = Path(f"{temp_db}{suffix}")
            if path.exists():
                path.unlink()

    static_data_path = write_static_app_data(
        dict(parquet_outputs)["grouped_finance_rows"],
        dict(parquet_outputs)["section_grouped_finance_rows"],
    )

    print(f"Wrote {len(parquet_outputs)} parquet tables to {PARQUET_DIR}")
    print(f"Wrote local app data to {static_data_path}")


if __name__ == "__main__":
    main()
