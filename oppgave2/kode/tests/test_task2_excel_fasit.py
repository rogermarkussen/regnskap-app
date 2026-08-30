from __future__ import annotations

import unittest
from pathlib import Path
import sys

import duckdb
from openpyxl import load_workbook

from scripts.prepare_data import (
    _cash_values_for_account,
    _empty_account_values,
    latest_complete_ledger_period,
)


CODE_ROOT = Path(__file__).resolve().parents[1]
ROOT = CODE_ROOT.parent
REPO_ROOT = ROOT.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from shared.data_contract import load_data_contract


CONTRACT = load_data_contract(REPO_ROOT)
FASIT_PATH = CONTRACT.path("fasit.dashboard_kpi")
LEDGER_PATH = CONTRACT.path("common.ledger")
BUDGET_HEADER_PATH = CONTRACT.path("common.budget_header")
BUDGET_VALUE_PATH = CONTRACT.path("common.budget_values")
TOLERANCE_TUSEN = 0.00001

SHEETS = {
    "Finansiering 154301": {
        "period_to": 202603,
        "actual_filter": "trim(dim_4) = '154301'",
        "budget_filter": "(h.dim_1 is null or trim(h.dim_1) not in ('212', '761'))",
    },
    "Finansiering 154345": {
        "period_to": 202604,
        "actual_filter": "trim(dim_4) = '154345'",
        "budget_filter": "trim(h.dim_1) = '212'",
    },
    "Fin 154322 og 045101": {
        "period_to": 202603,
        "actual_filter": "trim(dim_4) in ('154322', '045101')",
        "budget_filter": "trim(h.dim_1) = '761'",
    },
}


class Task2MissingCashTest(unittest.TestCase):
    def test_explicit_zero_is_preserved(self) -> None:
        budget, cash = _cash_values_for_account(
            {
                "5000": {
                    "kontant_budsjett_tusen": 0.0,
                    "kontant_tusen": 12.5,
                }
            },
            "5000",
            available=True,
        )

        self.assertEqual(budget, 0.0)
        self.assertEqual(cash, 12.5)

    def test_missing_cash_source_is_not_replaced_with_zero(self) -> None:
        self.assertEqual(
            _cash_values_for_account({}, "5000", available=True),
            (None, None),
        )
        self.assertEqual(
            _cash_values_for_account(
                {"5000": {"kontant_budsjett_tusen": 0.0, "kontant_tusen": 12.5}},
                "5000",
                available=False,
            ),
            (None, None),
        )

    def test_empty_grouping_account_has_no_invented_cash_values(self) -> None:
        values = _empty_account_values()

        self.assertEqual(values["hovedbok_tusen"], 0.0)
        self.assertIsNone(values["kontant_budsjett_tusen"])
        self.assertIsNone(values["kontant_tusen"])
        self.assertIsNone(values["kontant_avvik_tusen"])


class Task2LatestPeriodTest(unittest.TestCase):
    def test_latest_complete_period_matches_snapshot(self) -> None:
        expected = 202608 if CONTRACT.test_data_root else 202606
        self.assertEqual(latest_complete_ledger_period(), expected)


class Task2ReportPeriodTest(unittest.TestCase):
    def test_january_to_june_budget_is_sum_of_first_six_months(self) -> None:
        rows_path = CONTRACT.generated_dir("oppgave2") / "static-app" / "grouped_finance_rows.parquet"

        connection = duckdb.connect()
        try:
            columns = {
                row[0]
                for row in connection.execute(
                    f"describe select * from read_parquet('{rows_path.as_posix()}')"
                ).fetchall()
            }
            if "report_year" in columns:
                month_columns = [f"budsjett_{month:02d}_tusen" for month in range(1, 7)]
                period_filter = (
                    "report_year = 2026 and period_to = 202606 "
                    "and section_code = 'all'"
                )
            else:
                month_columns = [
                    f"budsjett_20260{month}_tusen" for month in range(1, 7)
                ]
                period_filter = "rapportperiode = 'p1_6'"
            month_sum = " + ".join(
                f"coalesce({column}, 0)" for column in month_columns
            )
            errors, checked = connection.execute(
                f"""
                select
                  count(*) filter (
                    where abs(virksomhet_budsjett_tusen - ({month_sum})) > 0.00001
                  ) as errors,
                  count(*) as checked
                from read_parquet('{rows_path.as_posix()}')
                where {period_filter}
                  and row_type in ('account', 'group', 'total')
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertGreater(checked, 0)
        self.assertEqual(errors, 0)

    def test_investment_columns_follow_154345_rule(self) -> None:
        rows_path = CONTRACT.generated_dir("oppgave2") / "static-app" / "grouped_finance_rows.parquet"
        connection = duckdb.connect()
        try:
            columns = {
                row[0]
                for row in connection.execute(
                    f"describe select * from read_parquet('{rows_path.as_posix()}')"
                ).fetchall()
            }
            if "report_year" in columns:
                totals_filter = "row_type = 'account'"
                join_keys = """
                  and investment.section_code = report.section_code
                  and investment.period_to = report.period_to
                  and investment.konto = report.konto
                """
            else:
                totals_filter = "radtekst = 'Driftskostnader'"
                join_keys = ""
            mismatches = connection.execute(
                f"""
                with totals as (
                  select *
                  from read_parquet('{rows_path.as_posix()}')
                  where {totals_filter}
                )
                select count(*)
                from totals report
                join totals investment
                  on investment.finansiering = '154345'
                 and investment.rapportperiode = report.rapportperiode
                 {join_keys}
                where report.finansiering in ('154345', 'alle')
                  and (
                    abs(report.investeringsbudsjett_tusen - investment.virksomhet_budsjett_tusen) > 0.00001
                    or abs(report.investeringsregnskap_tusen - investment.hovedbok_tusen) > 0.00001
                  )
                """
            ).fetchone()[0]
            non_investment_values = connection.execute(
                f"""
                select count(*)
                from read_parquet('{rows_path.as_posix()}')
                where finansiering in ('154301', '154322+045101')
                  and (
                    investeringsbudsjett_tusen is not null
                    or investeringsregnskap_tusen is not null
                  )
                """
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertEqual(mismatches, 0)
        self.assertEqual(non_investment_values, 0)

    def test_missing_budget_source_is_not_published_as_zero(self) -> None:
        rows_path = (
            CONTRACT.generated_dir("oppgave2")
            / "static-app"
            / "grouped_finance_rows.parquet"
        )
        connection = duckdb.connect()
        try:
            columns = {
                row[0]
                for row in connection.execute(
                    f"describe select * from read_parquet('{rows_path.as_posix()}')"
                ).fetchall()
            }
            if "report_year" not in columns:
                self.skipTest("Kontrollen gjelder den flerårige Parquet-rapporten")

            source_rows = connection.execute(
                f"""
                select count(*)
                from read_parquet('{BUDGET_HEADER_PATH.as_posix()}') h
                join read_parquet('{BUDGET_VALUE_PATH.as_posix()}') v using (trans_id)
                where h.version = '2026B'
                  and trim(h.account) = '5080'
                  and try_cast(v.period as integer) between 202601 and 202612
                """
            ).fetchone()[0]
            report_row = connection.execute(
                f"""
                select
                  virksomhet_budsjett_tusen,
                  aarets_budsjett_tusen,
                  avvik_tusen
                from read_parquet('{rows_path.as_posix()}')
                where section_code = 'all'
                  and finansiering = 'alle'
                  and rapportperiode = '202608'
                  and row_type = 'account'
                  and konto = '5080'
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(source_rows, 0)
        self.assertIsNotNone(report_row)
        self.assertEqual(report_row, (None, None, None))


def _account_rows(worksheet):
    for row_number in range(1, worksheet.max_row + 1):
        label = worksheet.cell(row_number, 1).value
        if not isinstance(label, str) or len(label) < 4 or not label[:4].isdigit():
            continue
        yield row_number, label[:4]


class Task2ExcelFasitTest(unittest.TestCase):
    """Excel is read only as an oracle; all tested values are calculated from Parquet."""

    @classmethod
    def setUpClass(cls) -> None:
        required = [FASIT_PATH, LEDGER_PATH, BUDGET_HEADER_PATH, BUDGET_VALUE_PATH]
        missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
        if missing:
            raise AssertionError(f"Mangler testgrunnlag: {', '.join(missing)}")

        cls.workbook = load_workbook(FASIT_PATH, read_only=True, data_only=True)
        cls.actual: dict[tuple[str, str], float] = {}
        cls.budget: dict[tuple[str, str, int], float] = {}
        connection = duckdb.connect()
        try:
            for sheet, rule in SHEETS.items():
                actual_rows = connection.execute(
                    f"""
                    select
                      lpad(trim(account), 4, '0') as konto,
                      sum(try_cast(amount as double)) / 1000 as belop
                    from read_parquet('{LEDGER_PATH.as_posix()}')
                    where {rule['actual_filter']}
                      and try_cast(period as integer) between 202601 and {rule['period_to']}
                    group by account
                    """
                ).fetchall()
                for account, amount in actual_rows:
                    cls.actual[(sheet, str(account))] = float(amount)

                budget_rows = connection.execute(
                    f"""
                    select
                      lpad(cast(h.account as varchar), 4, '0') as konto,
                      try_cast(v.period as integer) as periode,
                      sum(try_cast(v.amount as double)) / 1000 as belop
                    from read_parquet('{BUDGET_HEADER_PATH.as_posix()}') h
                    join read_parquet('{BUDGET_VALUE_PATH.as_posix()}') v using (trans_id)
                    where h.version = '2026B'
                      and {rule['budget_filter']}
                      and try_cast(v.period as integer) between 202601 and 202612
                    group by h.account, v.period
                    """
                ).fetchall()
                for account, period, amount in budget_rows:
                    cls.budget[(sheet, str(account), int(period))] = float(amount)
        finally:
            connection.close()

    @classmethod
    def tearDownClass(cls) -> None:
        cls.workbook.close()

    def assertCellEqual(
        self, *, sheet: str, cell: str, calculated: float, expected: float
    ) -> None:
        self.assertAlmostEqual(
            calculated,
            expected,
            delta=TOLERANCE_TUSEN,
            msg=(
                f"{sheet}!{cell}: beregnet {calculated:.9f}, "
                f"Excel {expected:.9f}, avvik {calculated - expected:.9f}"
            ),
        )

    def test_every_budget_cell_is_reproduced_from_parquet(self) -> None:
        comparisons = 0
        for sheet, rule in SHEETS.items():
            worksheet = self.workbook[sheet]
            for row_number, account in _account_rows(worksheet):
                monthly = {
                    period: self.budget.get((sheet, account, period), 0.0)
                    for period in range(202601, 202613)
                }
                for offset, period in enumerate(range(202601, 202613), start=10):
                    expected = worksheet.cell(row_number, offset).value
                    if isinstance(expected, (int, float)):
                        self.assertCellEqual(
                            sheet=sheet,
                            cell=f"{worksheet.cell(row_number, offset).coordinate}",
                            calculated=monthly[period],
                            expected=float(expected),
                        )
                        comparisons += 1

                annual = sum(monthly.values())
                period_budget = sum(
                    monthly[period]
                    for period in range(202601, int(rule["period_to"]) + 1)
                )
                for column, calculated in ((3, period_budget), (8, annual)):
                    expected = worksheet.cell(row_number, column).value
                    if isinstance(expected, (int, float)):
                        self.assertCellEqual(
                            sheet=sheet,
                            cell=worksheet.cell(row_number, column).coordinate,
                            calculated=calculated,
                            expected=float(expected),
                        )
                        comparisons += 1
        self.assertEqual(comparisons, 2282)

    def test_every_actual_and_derived_cell_is_reproduced(self) -> None:
        comparisons = 0
        for sheet, rule in SHEETS.items():
            worksheet = self.workbook[sheet]
            for row_number, account in _account_rows(worksheet):
                actual = self.actual.get((sheet, account), 0.0)
                period_budget = sum(
                    self.budget.get((sheet, account, period), 0.0)
                    for period in range(202601, int(rule["period_to"]) + 1)
                )
                annual = sum(
                    self.budget.get((sheet, account, period), 0.0)
                    for period in range(202601, 202613)
                )
                expected_actual = worksheet.cell(row_number, 4).value
                if isinstance(expected_actual, (int, float)):
                    self.assertCellEqual(
                        sheet=sheet,
                        cell=worksheet.cell(row_number, 4).coordinate,
                        calculated=actual,
                        expected=float(expected_actual),
                    )
                    comparisons += 1

                expected_deviation = worksheet.cell(row_number, 5).value
                if isinstance(expected_deviation, (int, float)):
                    source_deviation = period_budget - actual
                    self.assertCellEqual(
                        sheet=sheet,
                        cell=worksheet.cell(row_number, 5).coordinate,
                        calculated=source_deviation,
                        expected=float(expected_deviation),
                    )
                    comparisons += 1

                expected_consumption = worksheet.cell(row_number, 9).value
                if isinstance(expected_consumption, (int, float)) and annual:
                    source_consumption = actual / annual * 100
                    self.assertCellEqual(
                        sheet=sheet,
                        cell=worksheet.cell(row_number, 9).coordinate,
                        calculated=source_consumption,
                        expected=float(expected_consumption),
                    )
                    comparisons += 1
        self.assertEqual(comparisons, 400)

    def test_cash_cells_are_arithmetically_consistent(self) -> None:
        """No independent cash Parquet exists; do not claim source reconciliation."""
        checked = 0
        for sheet in SHEETS:
            worksheet = self.workbook[sheet]
            for row_number, _ in _account_rows(worksheet):
                budget = worksheet.cell(row_number, 23).value
                cash = worksheet.cell(row_number, 24).value
                deviation = worksheet.cell(row_number, 25).value
                if all(isinstance(value, (int, float)) for value in (budget, cash, deviation)):
                    self.assertCellEqual(
                        sheet=sheet,
                        cell=worksheet.cell(row_number, 25).coordinate,
                        calculated=float(budget) - float(cash),
                        expected=float(deviation),
                    )
                    checked += 1
        self.assertEqual(checked, 163)


if __name__ == "__main__":
    unittest.main()
