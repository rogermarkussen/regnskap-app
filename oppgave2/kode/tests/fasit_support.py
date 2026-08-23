from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


CODE_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = CODE_ROOT.parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from shared.data_contract import load_data_contract


CONTRACT = load_data_contract(REPO_ROOT)
FASIT_WORKBOOKS = {
    CONTRACT.path("fasit.account_grouping_154301"): ("154301", "Finansiering 154301"),
    CONTRACT.path("fasit.account_grouping_all"): ("alle", "Alle finansieringer"),
}


def _clean(value: object) -> object:
    if isinstance(value, str):
        value = value.replace("\xa0", " ").strip()
        return None if value == "" else value
    return value


def _account(value: object) -> tuple[str | None, str | None]:
    if not isinstance(value, str):
        return None, None
    match = re.match(r"^(\d{4})\s*-\s*(.+)$", value.strip())
    return (match.group(1), match.group(2).strip()) if match else (None, None)


def grouped_finance_fasit_rows_frame() -> pd.DataFrame:
    rows = []
    for workbook_path, (financing, financing_label) in FASIT_WORKBOOKS.items():
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook["Kontogruppering med tall"]
            current_main_group: str | None = None
            for excel_row, row in enumerate(
                worksheet.iter_rows(min_row=11, values_only=True), start=11
            ):
                values = [_clean(value) for value in row]
                label = values[0] if values else None
                if not label or label == "Exported 18.06.2026":
                    continue
                label = str(label)
                account_number, account_name = _account(label)
                numeric_values = [value for value in values[2:25] if isinstance(value, int | float)]
                if account_number:
                    row_type = "account"
                elif not numeric_values:
                    row_type = "section"
                    current_main_group = label
                elif label.lower().startswith("totale ") or label == "Driftskostnader":
                    row_type = "total"
                else:
                    row_type = "group"
                rows.append(
                    {
                        "finansiering": financing,
                        "finansiering_tekst": financing_label,
                        "excel_row": excel_row,
                        "hovedgruppe": current_main_group,
                        "row_type": row_type,
                        "radtekst": label,
                        "konto": account_number,
                        "konto_navn": account_name,
                        "virksomhet_budsjett_tusen": _number(values, 2),
                        "hovedbok_tusen": _number(values, 3),
                        "avvik_tusen": _number(values, 4),
                        "aarets_budsjett_tusen": _number(values, 7),
                        "forbruk_av_aarets_budsjett": _number(values, 8),
                        **{
                            f"budsjett_{period}_tusen": _number(values, index)
                            for index, period in enumerate(range(202601, 202613), start=9)
                        },
                        "kontant_budsjett_tusen": _number(values, 22),
                        "kontant_tusen": _number(values, 23),
                        "kontant_avvik_tusen": _number(values, 24),
                        "source_file": workbook_path.name,
                    }
                )
        finally:
            workbook.close()
    return pd.DataFrame(rows)


def _number(values: list[object], index: int) -> int | float | None:
    return values[index] if len(values) > index and isinstance(values[index], int | float) else None
