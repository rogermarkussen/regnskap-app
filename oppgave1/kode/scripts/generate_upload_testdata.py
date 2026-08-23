"""Lag syntetiske opplastingsfiler for manuell testing av oppgave 1."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "testdata-opplasting"
PERIODS = {"p1_3": 0.50, "p1_4": 0.75, "p1_6": 1.00}
PERIOD_LABELS = {"p1_3": "Jan-mar", "p1_4": "Jan-apr", "p1_6": "Jan-jun"}
END_PERIODS = {"p1_3": "202603", "p1_4": "202604", "p1_6": "202606"}

CALCULATION_RULES = {
    ("154301", "ADK"): "konto 6110–7834",
    ("154301", "Konsulent"): "konto 6700, 6710, 6720, 6730, 6731, 6732",
    ("154301", "Reise"): "konto 7100, 7130, 7131, 7150, 7190, 7199",
    ("154301", "Overtid"): "konto 5050, 5150",
    ("154301", "Lønnsandel"): "konto 5000–5999 / konto 5000–7834",
    ("154345", "Totalt regnskap vs budsjett"): "konto 6110–7834",
    ("154322+045101", "ADK"): "konto 6110–7834",
    ("154322+045101", "Testlab prosjekt 7114"): "konto 5000–7834, prosjekt 7114",
    ("154322+045101", "Lønnsandel"): "konto 5000–5999 / konto 5000–7834",
}

SIMPLE_ROWS = [
    ("154301", "ADK", 800, 1000, None, "Syntetisk testverdi"),
    ("154301", "Konsulent", 230, 300, None, "Syntetisk testverdi"),
    ("154301", "Reise", 125, 150, None, "Syntetisk testverdi"),
    ("154301", "Overtid", 90, 80, None, "Tester status over budsjett"),
    ("154301", "Lønnsandel", None, None, 0.42, "42 prosent"),
    ("154345", "Totalt regnskap vs budsjett", 500, 750, None, "Syntetisk testverdi"),
    ("154322+045101", "ADK", 650, 800, None, "Syntetisk testverdi"),
    ("154322+045101", "Testlab prosjekt 7114", 120, 200, None, "Syntetisk testverdi"),
    ("154322+045101", "Lønnsandel", None, None, 0.35, "35 prosent"),
]

HEADERS = [
    "rapportperiode",
    "finansiering",
    "tittel",
    "hovedbok_nok1000",
    "budsjett_nok1000",
    "prosentverdi",
    "kommentar",
]


def excel_records() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for period_key, factor in PERIODS.items():
        for financing, title, actual, budget, percentage, comment in SIMPLE_ROWS:
            rows.append({
                "rapportperiode": PERIOD_LABELS[period_key],
                "finansiering": financing,
                "tittel": title,
                "hovedbok_nok1000": None if actual is None else float(actual) * factor,
                "budsjett_nok1000": None if budget is None else float(budget) * factor,
                "prosentverdi": (
                    None if percentage is None else percentage - (1 - factor) * 0.08
                ),
                "kommentar": comment,
            })
    return rows


def write_excel() -> None:
    xlsx_path = OUTPUT / "excel" / "oppgave1_testdata.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.append(HEADERS)
    for record in excel_records():
        dashboard.append([record[field] for field in HEADERS])
    dashboard.freeze_panes = "A2"
    dashboard.auto_filter.ref = f"A1:G{dashboard.max_row}"
    for column, width in zip("ABCDEFG", (18, 20, 34, 22, 22, 18, 38), strict=True):
        dashboard.column_dimensions[column].width = width

    info = workbook.create_sheet("Info")
    info.append(["Testdata", "Alle verdier er syntetiske og skal kun brukes til opplastingstest."])
    info.append(["Prosent", "0,42 vises som 42 prosent i dashbordet."])
    info.append(["Perioder", "Filen har ni separate KPI-rader for hver av Jan-mar, Jan-apr og Jan-jun."])
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90
    workbook.save(xlsx_path)


def status_fields(actual: float, budget: float | None) -> tuple[float | None, str | None, str | None]:
    if budget in (None, 0):
        return None, None, None
    share = actual / budget
    if share > 1:
        return share, "danger", "Over budsjett"
    if share >= 0.85:
        return share, "warning", "Nær budsjett"
    return share, "ok", "Innenfor budsjett"


def calculated_records(scenario: str = "blandet") -> list[dict[str, object]]:
    if scenario not in {"blandet", "innenfor", "over_budsjett"}:
        raise ValueError(f"Ukjent testscenario: {scenario}")
    metric_names = {
        "Konsulent": "Konsulentkostnader",
        "Reise": "Reisekostnader",
        "Testlab prosjekt 7114": "Testlab",
        "Lønnsandel": "Lønnsandel av totale kostnader",
    }
    rows: list[dict[str, object]] = []
    for period_key, factor in PERIODS.items():
        for financing, title, actual, budget, percentage, comment in SIMPLE_ROWS:
            metric = metric_names.get(title, title)
            if percentage is not None:
                value = percentage - (1 - factor) * 0.08
                if scenario == "innenfor":
                    value = max(0.0, value - 0.05)
                elif scenario == "over_budsjett":
                    value = min(1.0, value + 0.10)
                details = [
                    {"label": "Lønnskostnader", "value": round(value * 1000, 6)},
                    {"label": "Totale kostnader", "value": 1000},
                    {"label": "Andel (%)", "value": round(value * 100, 6), "format": "pct"},
                ]
                rows.append({
                    "period_key": period_key,
                    "end_period": END_PERIODS[period_key],
                    "finansiering": financing,
                    "metric": metric,
                    "tittel": title,
                    "hovedbok_nok1000": value,
                    "budsjett_nok1000": None,
                    "budsjettandel": None,
                    "status": None,
                    "status_tekst": None,
                    "prosentverdi": value,
                    "gjenstaar_nok1000": -value,
                    "kommentar": (
                        comment
                        if scenario == "blandet"
                        else f"Syntetisk demo: {scenario.replace('_', ' ')}"
                    ),
                    "grunnlag_json": json.dumps(details, ensure_ascii=False),
                    "beregningsregel": CALCULATION_RULES[(financing, title)],
                    "regelversjon": "2026-08-06",
                    "budsjettversjon": "2026B",
                })
                continue

            scaled_actual = float(actual) * factor
            scaled_budget = float(budget) * factor
            if scenario == "innenfor":
                scaled_actual = scaled_budget * 0.70
            elif scenario == "over_budsjett":
                scaled_actual = scaled_budget * 1.10
            share, status, status_text = status_fields(scaled_actual, scaled_budget)
            rows.append({
                "period_key": period_key,
                "end_period": END_PERIODS[period_key],
                "finansiering": financing,
                "metric": metric,
                "tittel": title,
                "hovedbok_nok1000": scaled_actual,
                "budsjett_nok1000": scaled_budget,
                "budsjettandel": share,
                "status": status,
                "status_tekst": status_text,
                "prosentverdi": None,
                "gjenstaar_nok1000": scaled_budget - scaled_actual,
                "kommentar": (
                    comment
                    if scenario == "blandet"
                    else f"Syntetisk demo: {scenario.replace('_', ' ')}"
                ),
                "grunnlag_json": json.dumps(
                    [{"label": "Syntetisk grunnlag", "value": scaled_actual}],
                    ensure_ascii=False,
                ),
                "beregningsregel": CALCULATION_RULES[(financing, title)],
                "regelversjon": "2026-08-06",
                "budsjettversjon": "2026B",
            })
    return rows


def write_calculated_parquet() -> None:
    output = OUTPUT / "parquet" / "beregnet"
    output.mkdir(parents=True, exist_ok=True)
    scenarios = {
        "dashboard_kpi_testdata.parquet": "blandet",
        "dashboard_kpi_demo_innenfor.parquet": "innenfor",
        "dashboard_kpi_demo_over_budsjett.parquet": "over_budsjett",
    }
    for filename, scenario in scenarios.items():
        pd.DataFrame(calculated_records(scenario)).to_parquet(
            output / filename,
            index=False,
            compression="zstd",
        )


def write_operational_parquet() -> None:
    output = OUTPUT / "parquet" / "operative"
    output.mkdir(parents=True, exist_ok=True)
    months = [f"2026{month:02d}" for month in range(1, 7)]
    actual_specs = [
        ("154301", "6110", "", 100_000),
        ("154301", "6700", "", 30_000),
        ("154301", "7100", "", 15_000),
        ("154301", "5050", "", 10_000),
        ("154301", "5100", "", 80_000),
        ("154345", "6110", "", 60_000),
        ("154322", "6110", "", 70_000),
        ("154322", "6200", "7114", 20_000),
        ("154322", "5100", "", 50_000),
    ]
    actual_rows = [
        {"account": account, "dim_4": financing, "dim_2": project, "period": period, "amount": amount}
        for period in months
        for financing, account, project, amount in actual_specs
    ]
    pd.DataFrame(actual_rows).to_parquet(output / "agltransact.parquet", index=False, compression="zstd")

    budget_specs = [
        ("711", "6110", "", 120_000),
        ("711", "6700", "", 40_000),
        ("711", "7100", "", 20_000),
        ("711", "5050", "", 12_000),
        ("212", "6110", "", 80_000),
        ("761", "6110", "", 90_000),
        ("761", "6200", "7114", 30_000),
    ]
    headers = []
    values = []
    for index, (dim_1, account, project, amount) in enumerate(budget_specs, start=1):
        trans_id = f"TEST-{index:03d}"
        headers.append({
            "trans_id": trans_id,
            "account": account,
            "dim_1": dim_1,
            "dim_2": project,
            "version": "2026B",
        })
        values.extend(
            {"trans_id": trans_id, "period": period, "amount": amount}
            for period in months
        )
    pd.DataFrame(headers).to_parquet(output / "apltransact.parquet", index=False, compression="zstd")
    pd.DataFrame(values).to_parquet(output / "apltransactvalue.parquet", index=False, compression="zstd")


def main() -> None:
    write_excel()
    write_calculated_parquet()
    write_operational_parquet()
    print(f"Genererte opplastingstestdata i {OUTPUT}")


if __name__ == "__main__":
    main()
