"""Kontroller at de genererte manuelle opplastingsfilene har støttet format."""

from __future__ import annotations

import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
TESTDATA = ROOT / "testdata-opplasting"
HEADERS = [
    "rapportperiode",
    "finansiering",
    "tittel",
    "hovedbok_nok1000",
    "budsjett_nok1000",
    "prosentverdi",
    "kommentar",
]


class UploadTestdataTest(unittest.TestCase):
    def test_excel_har_dashboardark_og_riktig_format(self) -> None:
        workbook = load_workbook(
            TESTDATA / "excel" / "oppgave1_testdata.xlsx",
            read_only=True,
            data_only=True,
        )
        self.assertIn("Dashboard", workbook.sheetnames)
        rows = list(workbook["Dashboard"].iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), HEADERS)
        self.assertEqual(len(rows) - 1, 27)
        self.assertEqual({row[0] for row in rows[1:]}, {"Jan-mar", "Jan-apr", "Jan-jun"})

    def test_beregnet_parquet_har_ni_kpier_i_tre_perioder(self) -> None:
        filenames = [
            "dashboard_kpi_testdata.parquet",
            "dashboard_kpi_demo_innenfor.parquet",
            "dashboard_kpi_demo_over_budsjett.parquet",
        ]
        frames = {}
        for filename in filenames:
            frame = pd.read_parquet(TESTDATA / "parquet" / "beregnet" / filename)
            frames[filename] = frame
            self.assertEqual(len(frame), 27, filename)
            self.assertEqual(set(frame["period_key"]), {"p1_3", "p1_4", "p1_6"})
            self.assertEqual(set(frame["regelversjon"]), {"2026-08-06"})
            self.assertEqual(set(frame["budsjettversjon"]), {"2026B"})
            self.assertEqual(frame.groupby("period_key").size().to_dict(), {
                "p1_3": 9,
                "p1_4": 9,
                "p1_6": 9,
            })

        within = frames["dashboard_kpi_demo_innenfor.parquet"]
        over = frames["dashboard_kpi_demo_over_budsjett.parquet"]
        self.assertEqual(set(within["status"].dropna()), {"ok"})
        self.assertEqual(set(over["status"].dropna()), {"danger"})


if __name__ == "__main__":
    unittest.main()
